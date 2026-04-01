"""
allegati.py — Gestione allegati per progetto
Struttura disco: {tools_toa_folder}/allegati/{project_id}/{uid}_{filename}
Metadati:        {tools_toa_folder}/allegati/{project_id}/index.json
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from database.db_handler import carica_configurazione
from pathlib import Path
from datetime import datetime
import json, re, uuid

router = APIRouter(prefix="/api/allegati", tags=["allegati"])

TIPI_CONSENTITI = {
    ".xlsx": "setup_cam", ".xls": "setup_cam",
    ".pdf":  "documento",
    ".jpg":  "immagine",  ".jpeg": "immagine",
    ".png":  "immagine",  ".bmp": "immagine",
    ".dwg":  "disegno",   ".dxf": "disegno",
    ".txt":  "documento",
}

def _dir(config, pid):
    base = config.get("tools_toa_folder") or config.get("percorso_nc_base") or "."
    base = base.strip() or "."
    # Risolve path relativa rispetto alla directory di lavoro del processo
    p = Path(base).resolve() / "allegati" / pid
    return p

def _idx(d: Path) -> list:
    f = d / "index.json"
    if f.exists():
        try: return json.loads(f.read_text(encoding="utf-8"))
        except: pass
    return []

def _save_idx(d: Path, data: list):
    d.mkdir(parents=True, exist_ok=True)
    (d / "index.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def _safe(name: str) -> str:
    return re.sub(r"[^\w\s.\-]", "_", name)[:120]


@router.get("/{project_id}")
async def lista_allegati(project_id: str):
    config = carica_configurazione()
    d = _dir(config, project_id)
    result = []
    for a in _idx(d):
        p = d / a.get("path", "")
        if p.exists():
            result.append({**a, "size_kb": round(p.stat().st_size / 1024, 1)})
    return {"allegati": result, "n": len(result)}


@router.post("/{project_id}/upload")
async def upload_allegato(
    project_id: str,
    file: UploadFile = File(...),
    fase: str = Form(default=""),
    note: str = Form(default=""),
):
    config = carica_configurazione()
    d = _dir(config, project_id)
    d.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename or "").suffix.lower()
    if ext not in TIPI_CONSENTITI:
        raise HTTPException(400, f"Tipo non supportato: {ext}")
    tipo = TIPI_CONSENTITI[ext]
    uid  = uuid.uuid4().hex[:8]
    dest_fn = f"{uid}_{_safe(file.filename or 'file')}"
    dest    = d / dest_fn
    data    = await file.read()
    dest.write_bytes(data)
    entry = {
        "id":          uid,
        "filename":    file.filename,
        "path":        dest_fn,
        "tipo":        tipo,
        "fase":        fase.strip(),
        "note":        note.strip(),
        "data_upload": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "size_bytes":  len(data),
    }
    if tipo == "setup_cam":
        try: entry["meta"] = _parse_setup_cam(dest)["meta"] | \
                              {"statistiche": _parse_setup_cam(dest)["statistiche"]}
        except: pass
    idx = _idx(d)
    idx.append(entry)
    _save_idx(d, idx)
    return {"ok": True, "allegato": {**entry, "size_kb": round(len(data)/1024,1)}}


@router.get("/{project_id}/file/{allegato_id}")
async def scarica_allegato(project_id: str, allegato_id: str):
    config = carica_configurazione()
    d = _dir(config, project_id)
    idx_data = _idx(d)
    entry = next((a for a in idx_data if a["id"]==allegato_id), None)
    if not entry:
        raise HTTPException(404, f"Allegato {allegato_id} non trovato in {d}")
    p = d / entry["path"]
    if not p.exists():
        raise HTTPException(404, f"File fisico non trovato: {p}")
    return FileResponse(str(p), filename=entry["filename"], media_type="application/octet-stream")


@router.get("/{project_id}/preview/{allegato_id}")
async def preview_allegato(project_id: str, allegato_id: str):
    config = carica_configurazione()
    d = _dir(config, project_id)
    entry = next((a for a in _idx(d) if a["id"]==allegato_id), None)
    if not entry:
        raise HTTPException(404, f"Allegato {allegato_id} non trovato in {d}")
    p = d / entry["path"]
    if not p.exists():
        raise HTTPException(404, f"File fisico non trovato: {p}")
    ext = Path(entry["filename"]).suffix.lower()
    mt = {".jpg":"image/jpeg",".jpeg":"image/jpeg",".png":"image/png",
          ".bmp":"image/bmp",".pdf":"application/pdf",
          ".xlsx":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    return FileResponse(str(p), media_type=mt.get(ext,"application/octet-stream"))


@router.delete("/{project_id}/file/{allegato_id}")
async def elimina_allegato(project_id: str, allegato_id: str):
    config = carica_configurazione()
    d = _dir(config, project_id)
    idx = _idx(d)
    entry = next((a for a in idx if a["id"]==allegato_id), None)
    if not entry: raise HTTPException(404, "Non trovato")
    p = d / entry["path"]
    if p.exists(): p.unlink()
    _save_idx(d, [a for a in idx if a["id"]!=allegato_id])
    return {"ok": True}


@router.get("/{project_id}/setup-cam/{allegato_id}")
async def get_setup_cam(project_id: str, allegato_id: str):
    config = carica_configurazione()
    d = _dir(config, project_id)
    entry = next((a for a in _idx(d) if a["id"]==allegato_id), None)
    if not entry: raise HTTPException(404, "Non trovato")
    if entry.get("tipo") != "setup_cam": raise HTTPException(400, "Non è un Setup CAM")
    p = d / entry["path"]
    if not p.exists(): raise HTTPException(404, "File mancante")
    try:
        return {"ok": True, "allegato": entry, "data": _parse_setup_cam(p)}
    except Exception as e:
        raise HTTPException(500, f"Errore parsing: {e}")


def _parse_setup_cam(path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]

    # Metadati
    meta = {}
    LABELS = {"Comm. e Pos.":"commessa","Program name":"program_name",
               "Programmatore":"programmatore","Tipo di Lavorazione":"tipo_lavorazione",
               "Terna Di Lavorazione":"fase"}
    for row in rows[:14]:
        row_str = " ".join(str(v) for v in row if v)
        for lbl, key in LABELS.items():
            if lbl in row_str and key not in meta:
                for v in row:
                    s = str(v).strip() if v else ""
                    if s and lbl not in s and len(s) > 1:
                        meta[key] = s; break
        if "Data" in row_str and "/" in row_str and "data" not in meta:
            for v in row:
                if v and "/" in str(v): meta["data"] = str(v).strip(); break

    # Operazioni
    ops, in_ops = [], False
    for row in rows:
        if str(row[0] or "").strip() == "Proc.": in_ops = True; continue
        if in_ops and isinstance(row[0], (int,float)) and row[0] == int(row[0] or 0):
            commento = str(row[10] or "")
            m = re.search(r"Tempo:(\d+:\d+:\d+)", commento)
            ops.append({
                "proc": int(row[0]),
                "alias": str(row[1] or "").strip(),
                "utensile": str(row[2] or "").strip(),
                "lung_ut": row[4], "refrig": str(row[5] or ""),
                "offset": row[7], "z_step": row[8],
                "z_min": round(float(row[9]),3) if row[9] else None,
                "tempo": m.group(1) if m else "",
                "s": row[12], "f": row[13],
                "axis": str(row[14] or ""), "sr": str(row[15] or ""),
            })
        if str(row[0] or "") == "Elenco Utensili": break

    # Utensili
    uts, in_ut = [], False
    for row in rows:
        if str(row[0] or "") == "Elenco Utensili": in_ut = True; continue
        if in_ut and str(row[0] or "").strip() == "Numero": continue
        if in_ut and row[0] and "T" in str(row[0]):
            uts.append({"t_number":str(row[0]).strip(),"nome":str(row[1] or "").strip(),
                        "alias":str(row[6] or "").strip(),"diametro":str(row[8] or "").strip(),
                        "raggio":str(row[9] or "").strip(),"fuori_pinza":row[10],
                        "porta_utensile":str(row[11] or "").strip()})
        if in_ut and row[1] == "Motion Limits": break

    # Statistiche
    stats = {}
    for row in rows:
        k = str(row[6] or "")
        if k == "No. di operazioni": stats["n_operazioni"] = row[8]
        elif k == "No. di utensili": stats["n_utensili"] = row[8]
        elif k == "Total air time": stats["air_time"] = str(row[8] or "")
        elif k == "Total feed time": stats["feed_time"] = str(row[8] or "")
        elif k == "Total time": stats["total_time"] = str(row[8] or "")

    return {"meta": meta, "operazioni": ops, "utensili": uts, "statistiche": stats}


@router.get("/{project_id}/debug")
async def debug_allegati(project_id: str):
    """Mostra dove l'endpoint cerca i file — utile per diagnostica."""
    config = carica_configurazione()
    d = _dir(config, project_id)
    idx_data = _idx(d)
    files_check = []
    for a in idx_data:
        p = d / a.get("path","")
        files_check.append({
            "id": a["id"],
            "filename": a["filename"],
            "path_assoluto": str(p),
            "esiste": p.exists(),
        })
    return {
        "tools_toa_folder": config.get("tools_toa_folder"),
        "percorso_nc_base": config.get("percorso_nc_base"),
        "directory_allegati": str(d),
        "directory_esiste": d.exists(),
        "n_allegati_index": len(idx_data),
        "files": files_check,
    }
