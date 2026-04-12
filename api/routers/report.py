"""
api/routers/report.py
=====================
Sistema di raccolta e reportistica tempi lavorazione.

Struttura lavorazioni_log.json:
{
  "sessioni": [
    {
      "id": "uuid",
      "data": "2026-03-30",
      "progetto": "4297_0008",
      "pallet": 6,
      "inizio": "2026-03-30T10:16:46",
      "fine": "2026-03-30T14:30:00",
      "durata_sec": 15194,
      "programmi": [
        {
          "filename": "4297_0008_01_023.MPF",
          "commessa": "4297", "posizione": "0008", "fase": "01", "seq": "023",
          "inizio": "2026-03-30T10:16:54",
          "fine": "2026-03-30T10:23:11",
          "durata_sec": 377,
          "utensile": "FFPI42R0.8L114",
          "t_number": 3524
        }
      ],
      "gap_sec": 120,           # tempo fermo tra programmi
      "utensili": {             # accumulo ore per utensile
        "FFPI42R0.8L114": 1240,
        "RENISHAW": 45
      }
    }
  ]
}
"""

import json
from api.constants import (
    ORE_TURNO_SEC, OEE_QUALITA_DEFAULT, LOG_RETENTION_DAYS,
    CICLI_FINESTRA, CICLI_MIN_ANOMALIA, CICLI_ANOMALIA_SIGMA,
    TEMPI_CICLO_MIN_DURATA, TEMPI_CICLO_MAX_DURATA,
)
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
from fastapi import Body, APIRouter, Query
from fastapi.responses import FileResponse
from database.db_handler import carica_configurazione

router = APIRouter()

# ── Helpers percorso ──────────────────────────────────────────────────────────

def _log_path(config: dict) -> Path:
    base = (config.get("tools_toa_folder") or
            config.get("percorso_nc_base") or ".")
    return Path(base) / "lavorazioni_log.json"

_log_cache: dict = {"data": None, "ts": 0.0, "mtime": 0.0}
_LOG_CACHE_TTL = 4.0  # secondi — inferiore al poll interval (5s) ma evita letture multiple nella stessa request

def _load_log(config: dict) -> dict:
    import time as _time
    p = _log_path(config)
    now = _time.monotonic()

    # Cache valida se: non scaduta E file non modificato
    try:
        mtime = p.stat().st_mtime if p.exists() else 0.0
    except OSError:
        mtime = 0.0

    if (_log_cache["data"] is not None
            and (now - _log_cache["ts"]) < _LOG_CACHE_TTL
            and mtime == _log_cache["mtime"]):
        return _log_cache["data"]

    if p.exists():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                raise ValueError("Root non è un oggetto JSON")
            if not isinstance(data.get("sessioni"), list):
                data["sessioni"] = []
            if not isinstance(data.get("stato_corrente"), dict):
                data["stato_corrente"] = {}
            _log_cache["data"]  = data
            _log_cache["ts"]    = now
            _log_cache["mtime"] = mtime
            return data
        except (json.JSONDecodeError, ValueError) as e:
            from utils.logger import get_logger as _get_log
            _get_log("routers.report").error(
                f"lavorazioni_log.json corrotto: {e} — uso struttura vuota"
            )
        except Exception:
            pass
    return {"sessioni": [], "stato_corrente": {}}

def _save_log(config: dict, data: dict):
    """
    Scrittura atomica con pruning automatico.
    Mantiene solo le sessioni degli ultimi 90 giorni nel file principale.
    Le sessioni più vecchie vengono archiviate in lavorazioni_YYYY.json
    per preservare lo storico senza appesantire il file operativo.
    """
    # Invalida la cache — il file sta per cambiare
    _log_cache["data"] = None
    _log_cache["ts"]   = 0.0
    from datetime import timedelta as _td
    p = _log_path(config)
    p.parent.mkdir(parents=True, exist_ok=True)

    # Pruning: separa sessioni recenti da quelle da archiviare
    RETENTION_DAYS = LOG_RETENTION_DAYS
    cutoff = (datetime.now() - _td(days=RETENTION_DAYS)).strftime("%Y-%m-%d")
    sessioni_all    = data.get("sessioni", [])
    sessioni_recent = [s for s in sessioni_all if (s.get("data") or "9999") >= cutoff]
    sessioni_old    = [s for s in sessioni_all if (s.get("data") or "9999") < cutoff]

    # Archivia sessioni vecchie per anno (non blocca se fallisce)
    if sessioni_old:
        try:
            anni = {}
            for s in sessioni_old:
                anno = (s.get("data") or "0000")[:4]
                anni.setdefault(anno, []).append(s)
            for anno, sess_anno in anni.items():
                arch_path = p.parent / f"lavorazioni_{anno}.json"
                try:
                    arch_existing = json.loads(arch_path.read_text(encoding="utf-8")) if arch_path.exists() else {"sessioni": []}
                except Exception:
                    arch_existing = {"sessioni": []}
                # Merge evitando duplicati per id
                ids_esistenti = {s.get("id") for s in arch_existing.get("sessioni", [])}
                nuove = [s for s in sess_anno if s.get("id") not in ids_esistenti]
                arch_existing["sessioni"].extend(nuove)
                arch_tmp = arch_path.with_suffix(".tmp")
                arch_tmp.write_text(json.dumps(arch_existing, ensure_ascii=False, indent=2), encoding="utf-8")
                arch_tmp.replace(arch_path)
        except Exception:
            pass  # archivio non critico — non blocca il salvataggio principale

    data["sessioni"] = sessioni_recent

    # Scrittura atomica del file principale
    tmp = p.with_suffix(".tmp")
    try:
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(p)
    except Exception:
        try: tmp.unlink(missing_ok=True)
        except Exception: pass
        raise

def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def _now_str() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M")

def _parse_nome(filename: str) -> dict:
    """4297_0008_01_023.MPF → commessa/posizione/fase/seq"""
    n = filename.upper().replace(".MPF", "")
    p = n.split("_")
    return {
        "commessa":  p[0] if len(p) > 0 else n,
        "posizione": p[1] if len(p) > 1 else "",
        "fase":      p[2] if len(p) > 2 else "",
        "seq":       p[3] if len(p) > 3 else "",
    }

def _durata_str(sec: int) -> str:
    if sec is None: return "—"
    h, r = divmod(int(sec), 3600)
    m, s = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"

# ── Aggiornamento da log OpcUa (chiamato da macchina_live) ───────────────────

def aggiorna_da_log(
    programma_attivo: Optional[str],
    stato_pgm: int,
    pallet_num: Optional[int],
    progetto_nome: Optional[str],
    utensile: Optional[str],
    t_number: Optional[str],
    config: dict,
    override_feed: Optional[int] = None,
    override_mandrino: Optional[int] = None,
    stop_type: Optional[str] = None,   # "reset" | "stop" | "search" | None
    progetto_id: Optional[str] = None,
):
    """
    Chiamato da aggiorna-stati-da-log ogni 5 secondi.
    Registra inizio/fine programmi e accumula tempi utensili.
    Registra anche periodi con override ridotto per diagnostica.
    """
    data = _load_log(config)
    sc   = data.setdefault("stato_corrente", {})
    now  = _now_iso()
    dirty = False
    try:

        # ── Sanità: chiudi sessioni orfane da riavvio backend ─────────────────────
        # Se il backend è stato riavviato mentre la macchina girava, la sessione
        # corrente è rimasta aperta (fine=None). Se ora la macchina è ferma,
        # chiudiamo la sessione orfana usando l'ultimo_tick come timestamp di fine.
        if stato_pgm in (0, 5) and sc.get("sessione_id"):
            sess_orfana = _find_sess(data, sc["sessione_id"])
            if sess_orfana and sess_orfana.get("fine") is None:
                fine_orfana = sc.get("ultimo_tick") or now
                _chiudi_sessione(data, sc, fine_orfana)
                # Preserva fermo giornaliero prima di clear
                _fermo_salvato = sc.get("fermo_sec_giornaliero", 0)
                _fermo_data    = sc.get("fermo_data", "")
                sc.clear()
                if _fermo_salvato and _fermo_data == now[:10]:
                    sc["fermo_sec_giornaliero"] = _fermo_salvato
                    sc["fermo_data"]            = _fermo_data
                dirty = True

        # ── Macchina FERMA ────────────────────────────────────────────────────────
        # GRACE PERIOD: se la macchina si ferma (M0, E-Stop, pausa operatore)
        # non chiudiamo immediatamente la sessione. Aspettiamo fino a 15 minuti.
        # Se il MEDESIMO programma riprende entro quel tempo → continuazione.
        # Se passa più tempo o cambia programma → chiudi normalmente.
        #
        # TEMPO FERMO: accumulato in sc["fermo_sec_giornaliero"] ad ogni tick
        # con stato=5/0, indipendentemente dall'esistenza di una sessione aperta.
        # Questo copre anche il caso "macchina ferma dall'inizio del turno".
        GRACE_SEC = 900  # 15 minuti
        if stato_pgm in (0, 5):
            # ── Accumulo tempo fermo giornaliero (sempre, con o senza sessione) ──
            ultimo_fermo = sc.get("ultimo_tick_fermo")
            today = now[:10]

            if not ultimo_fermo or ultimo_fermo[:10] != today:
                # Primo tick fermo della giornata corrente.
                # Se ultimo_tick_fermo era ieri → backend era su ieri e macchina
                # era ferma dalla mezzanotte: recupera fermo da mezzanotte a ora.
                # Se ultimo_tick_fermo è null → backend appena avviato: non recuperare,
                # non sappiamo da quando è ferma.
                if sc.get("fermo_data") != today:
                    sc["fermo_sec_giornaliero"] = 0
                    sc["fermo_data"] = today
                    dirty = True
                if ultimo_fermo and ultimo_fermo[:10] < today:
                    # Era ferma ieri — recupera da mezzanotte a ora
                    try:
                        ts_now     = datetime.fromisoformat(now)
                        mezzanotte = datetime.fromisoformat(today + "T00:00:00")
                        fermo_recuperato = max(0, int((ts_now - mezzanotte).total_seconds()))
                        if sc.get("fermo_sec_giornaliero", 0) < fermo_recuperato:
                            sc["fermo_sec_giornaliero"] = fermo_recuperato
                        dirty = True
                    except Exception:
                        pass
                # Se ultimo_fermo è null: primo avvio backend, non recuperiamo nulla
            if ultimo_fermo:
                try:
                    delta_fermo = int((datetime.fromisoformat(now) -
                                       datetime.fromisoformat(ultimo_fermo)).total_seconds())
                    if 0 < delta_fermo <= 120:
                        # Tick normale — accumula delta
                        if sc.get("fermo_data") != today:
                            sc["fermo_sec_giornaliero"] = 0
                            sc["fermo_data"] = today
                        sc["fermo_sec_giornaliero"] = sc.get("fermo_sec_giornaliero", 0) + delta_fermo
                        dirty = True
                    elif delta_fermo > 120:
                        # Gap lungo (backend era giù o tick mancato).
                        # La macchina era già ferma (stato 0/5) prima del gap
                        # → recupera il fermo reale limitato alla giornata corrente.
                        # Non contiamo fermo di ieri nel contatore di oggi.
                        try:
                            ts_ultimo = datetime.fromisoformat(ultimo_fermo)
                            ts_now    = datetime.fromisoformat(now)
                            # Inizio della giornata corrente
                            inizio_oggi = datetime.fromisoformat(today + "T00:00:00")
                            # Il fermo recuperabile è solo da mezzanotte a ora
                            ts_da = max(ts_ultimo, inizio_oggi)
                            fermo_recuperato = max(0, int((ts_now - ts_da).total_seconds()))
                            if sc.get("fermo_data") != today:
                                sc["fermo_sec_giornaliero"] = 0
                                sc["fermo_data"] = today
                            sc["fermo_sec_giornaliero"] = sc.get("fermo_sec_giornaliero", 0) + fermo_recuperato
                            dirty = True
                        except Exception:
                            pass
                except Exception:
                    pass
            sc["ultimo_tick_fermo"] = now
            # Traccia inizio fermo globale (per heatmap/tabella fermi)
            if not sc.get("fermo_globale_inizio"):
                sc["fermo_globale_inizio"] = now
            dirty = True

            if sc.get("in_esecuzione"):
                # Registra tipo di fermo e timestamp pausa
                sess_corrente2 = _find_sess(data, sc.get("sessione_id", ""))
                if sess_corrente2 and stop_type:
                    sess_corrente2.setdefault("fermi", [])
                    sess_corrente2["fermi"].append({
                        "ts":   now,
                        "tipo": stop_type,
                    })
                    if stop_type == "reset":
                        sess_corrente2["n_fermi_anomali"] = sess_corrente2.get("n_fermi_anomali", 0) + 1
                    elif stop_type == "stop":
                        sess_corrente2["n_fermi_pianificati"] = sess_corrente2.get("n_fermi_pianificati", 0) + 1

                # Entra in grace period invece di chiudere subito
                sc["in_pausa"]      = True
                sc["pausa_inizio"]  = now
                sc["in_esecuzione"] = False
                dirty = True

            elif sc.get("in_pausa"):
                # Già in pausa — accumula tempo pausa in gap_sec della sessione
                # e controlla se il grace period è scaduto
                try:
                    pausa_sec = int((datetime.fromisoformat(now) -
                                     datetime.fromisoformat(sc["pausa_inizio"])).total_seconds())
                except Exception:
                    pausa_sec = GRACE_SEC + 1

                # Accumula il tempo di pausa direttamente in gap_sec della sessione
                if sc.get("sessione_id"):
                    sess_p = _find_sess(data, sc["sessione_id"])
                    if sess_p and ultimo_fermo:
                        try:
                            delta_p = int((datetime.fromisoformat(now) -
                                           datetime.fromisoformat(ultimo_fermo)).total_seconds())
                            if 0 < delta_p <= 120:
                                sess_p["gap_sec"] = sess_p.get("gap_sec", 0) + delta_p
                                dirty = True
                        except Exception:
                            pass

                if pausa_sec > GRACE_SEC:
                    # Grace scaduto → chiudi definitivamente
                    _chiudi_sessione(data, sc, sc.get("pausa_inizio") or now)
                    _fs, _fd = sc.get("fermo_sec_giornaliero",0), sc.get("fermo_data","")
                    sc.clear()
                    if _fs and _fd == now[:10]: sc["fermo_sec_giornaliero"]=_fs; sc["fermo_data"]=_fd
                    dirty = True

        # ── Stato 2 senza programma utente = fine ciclo / USERENDPROG ──────────────
        # progStatus=2 con programma di sistema (USERENDPROG, _N_CMA_DIR) significa
        # che il ciclo è finito e il CNC sta eseguendo routine di fine programma.
        # Trattiamo come "macchina ferma" per chiudere la sessione corrente.
        if stato_pgm == 2 and not programma_attivo:  # USERENDPROG o fine ciclo
            if sc.get("in_esecuzione"):
                sc["in_pausa"]      = True
                sc["pausa_inizio"]  = now
                sc["in_esecuzione"] = False
                dirty = True
            elif sc.get("in_pausa"):
                try:
                    pausa_sec = int((datetime.fromisoformat(now) -
                                     datetime.fromisoformat(sc["pausa_inizio"])).total_seconds())
                except Exception:
                    pausa_sec = GRACE_SEC + 1
                if pausa_sec > GRACE_SEC:
                    _chiudi_sessione(data, sc, sc.get("pausa_inizio") or now)
                    _fs, _fd = sc.get("fermo_sec_giornaliero",0), sc.get("fermo_data","")
                    sc.clear()
                    if _fs and _fd == now[:10]:
                        sc["fermo_sec_giornaliero"] = _fs
                        sc["fermo_data"]            = _fd
                    dirty = True

        # ── Macchina IN ESECUZIONE ────────────────────────────────────────────────
        # stato 2 con programma utente = posizionamento/ricerca blocco durante lavorazione
        # trattato come in esecuzione per aggiornare programma_corrente e ultimo_tick
        if stato_pgm in (1, 2, 3) and programma_attivo:
            # Filtra programmi di sistema Sinumerik e MAIN generati da AnalisiNC
            _FILTRI_SISTEMA = (
                "_N_CMA_DIR", "_N_CST_DIR", "_N_SYF_DIR",
                "_N_MPF_DIR", "_SPF", "BPOSAXIS", "TMPCYC",
                "0_MAIN_",     # MAIN generato da AnalisiNC
                "PALLET5",     # File pallet Siemens
                "ROTTURA",     # Ciclo rottura utensile
                "LASEREIN",    # Misura laser
                "BLMEAS",      # Block measurement
                "FINE_PALLET", # Fine pallet
                "BAX2START",   # Routine avvio
            )
            for f in _FILTRI_SISTEMA:
                if f in programma_attivo.upper():
                    programma_attivo = None
                    break

        if stato_pgm in (1, 2, 3) and programma_attivo:
            # ── Normalizzazione suffix numerico ────────────────────────────────
            # Il Siemens può registrare lo stesso programma con suffix diverso:
            # 4297_0006_001 durante EXTCALL dal MAIN
            # 4297_0006_801 se il MAIN usava numerazione N800+ nei blocchi
            # o se esiste una copia rinominata sulla share.
            # Normalizziamo: se il programma corrente in sessione ha lo stesso
            # prefisso (tutto tranne l'ultimo token) consideriamoli identici.
            prev_prog = sc.get("programma_corrente")
            if prev_prog and prev_prog.upper() != programma_attivo.upper():
                def _prefix(fname):
                    """Ritorna il prefisso senza suffix numerico Siemens (≥3 cifre, ≥800)."""
                    base = (fname or "").upper().replace(".MPF", "")
                    parts = base.split("_")
                    # Rimuovi suffix solo se è >= 800 (pattern Siemens: _801, _900, ecc.)
                    # NON rimuovere suffix sequenziali normali: 001, 002, 003...
                    if parts and parts[-1].isdigit() and len(parts[-1]) >= 3 and int(parts[-1]) >= 800:
                        return "_".join(parts[:-1])
                    return base

                if _prefix(prev_prog) == _prefix(programma_attivo) and _prefix(prev_prog) != prev_prog.upper().replace(".MPF",""):
                    # Stesso programma con suffix Siemens diverso → tratta come continuazione
                    programma_attivo = prev_prog  # usa il nome già registrato

            # Quando la macchina riparte, azzera il tracker del fermo
            # Registra l'intervallo di fermo globale (per heatmap/tabella)
            fermo_glob_inizio = sc.pop("fermo_globale_inizio", None)
            if fermo_glob_inizio:
                try:
                    _dur = int((datetime.fromisoformat(now) -
                                datetime.fromisoformat(fermo_glob_inizio)).total_seconds())
                    if _dur >= 60:  # solo fermi > 1 minuto
                        # La data del fermo è quella dell'INIZIO, non della fine.
                        # Se attraversa mezzanotte, tronca la durata alla mezzanotte
                        # e registra solo la parte nella giornata di inizio.
                        _data_inizio = fermo_glob_inizio[:10]
                        _data_fine   = now[:10]
                        if _data_inizio != _data_fine:
                            # Fermo a cavallo di mezzanotte — tronca a mezzanotte
                            try:
                                _mezzanotte = datetime.fromisoformat(_data_fine + "T00:00:00")
                                _ts_inizio  = datetime.fromisoformat(fermo_glob_inizio)
                                _dur_troncato = int((_mezzanotte - _ts_inizio).total_seconds())
                                if _dur_troncato >= 60:
                                    data.setdefault("fermi_globali", []).append({
                                        "inizio": fermo_glob_inizio,
                                        "fine":   _data_fine + "T00:00:00",
                                        "durata_sec": _dur_troncato,
                                        "data":   _data_inizio,
                                    })
                                # Aggiungi anche la parte di oggi (da mezzanotte a ora)
                                _dur_oggi = int((datetime.fromisoformat(now) - _mezzanotte).total_seconds())
                                if _dur_oggi >= 60:
                                    data.setdefault("fermi_globali", []).append({
                                        "inizio": _data_fine + "T00:00:00",
                                        "fine":   now,
                                        "durata_sec": _dur_oggi,
                                        "data":   _data_fine,
                                    })
                            except Exception:
                                # Fallback: usa dati originali con data inizio
                                data.setdefault("fermi_globali", []).append({
                                    "inizio": fermo_glob_inizio,
                                    "fine":   now,
                                    "durata_sec": _dur,
                                    "data":   _data_inizio,
                                })
                        else:
                            data.setdefault("fermi_globali", []).append({
                                "inizio": fermo_glob_inizio,
                                "fine":   now,
                                "durata_sec": _dur,
                                "data":   _data_inizio,
                            })
                        # Mantieni solo ultimi 500 fermi globali
                        if len(data["fermi_globali"]) > 500:
                            data["fermi_globali"] = data["fermi_globali"][-500:]
                        dirty = True
                except Exception:
                    pass
            sc.pop("ultimo_tick_fermo", None)

            # ── Ripresa dopo grace period (stesso programma) ──────────────────────
            # Se eravamo in pausa e il MEDESIMO programma riprende entro 15min:
            # non è un nuovo ciclo, è una continuazione. Non azzerare il timer.
            if sc.get("in_pausa") and prev_prog == programma_attivo:
                try:
                    pausa_sec = int((datetime.fromisoformat(now) -
                                     datetime.fromisoformat(sc["pausa_inizio"])).total_seconds())
                except Exception:
                    pausa_sec = 0
                if pausa_sec <= GRACE_SEC:
                    # Continuazione — ripristina esecuzione senza azzerare inizio
                    sc["in_pausa"]      = False
                    sc["pausa_inizio"]  = None
                    sc["in_esecuzione"] = True
                    sc["ultimo_tick"]   = now
                    dirty = True
                    # Salta il resto (non è un nuovo programma)
                    if dirty:
                        pass  # continua al blocco accumulo tick sotto
                else:
                    # Grace scaduto prima della ripresa — chiudi e tratta come nuovo
                    _chiudi_sessione(data, sc, sc.get("pausa_inizio") or now)
                    _fs, _fd = sc.get("fermo_sec_giornaliero",0), sc.get("fermo_data","")
                    sc.clear()
                    if _fs and _fd == now[:10]: sc["fermo_sec_giornaliero"]=_fs; sc["fermo_data"]=_fd
                    prev_prog = None
            elif sc.get("in_pausa"):
                # Programma diverso dopo pausa → chiudi la sessione precedente
                _chiudi_sessione(data, sc, sc.get("pausa_inizio") or now)
                _fs, _fd = sc.get("fermo_sec_giornaliero",0), sc.get("fermo_data","")
                sc.clear()
                if _fs and _fd == now[:10]: sc["fermo_sec_giornaliero"]=_fs; sc["fermo_data"]=_fd
                prev_prog = None

            # ── Cambio data (mezzanotte): chiudi sessione del giorno precedente ──
            sess_corrente = _find_sess(data, sc["sessione_id"]) if sc.get("sessione_id") else None
            if sess_corrente and sess_corrente.get("data") != now[:10]:
                # Chiudi la sessione di ieri con il timestamp di mezzanotte
                mezzanotte = now[:10] + "T00:00:00"
                _chiudi_sessione(data, sc, mezzanotte)
                sc.clear()  # A mezzanotte il fermo si azzera — nuovo giorno
                prev_prog = None  # Forza apertura nuova sessione sotto

            # Prima volta o cambio programma
            _log_rpt = __import__('logging').getLogger("routers.report")
            _log_rpt.info(f"aggiorna_da_log: stato={stato_pgm} prog={programma_attivo} prev={prev_prog} sc_sess={sc.get('sessione_id')} in_ex={sc.get('in_esecuzione')}")
            if prev_prog != programma_attivo:

                # Chiudi programma precedente
                if prev_prog and sc.get("sessione_id"):
                    _chiudi_programma(data, sc, now)

                # Avvia nuova sessione se non esiste
                if not sc.get("sessione_id"):
                    sess_id = str(uuid.uuid4())[:8]

                    # ── Contesto inizio sessione (per ML) ──────────────────
                    # Snapshot della situazione al momento esatto dell'avvio
                    contesto_avvio = {}
                    try:
                        from api.routers.tool_history import _load_history as _th_load
                        from database.db_handler import carica_configurazione as _cfg_ctx
                        from pathlib import Path as _Path
                        import json as _json
                        _cfg_ctx_data = _cfg_ctx()
                        _tools_path = _Path(
                            _cfg_ctx_data.get("tools_toa_folder") or "."
                        ) / "tools_machine.json"
                        if _tools_path.exists():
                            _tm = _json.loads(_tools_path.read_text(encoding="utf-8"))
                            _tools = list((_tm.get("tools") or {}).values())
                            contesto_avvio["utensili_montati"]   = len(_tools)
                            # Usa soglie ML per utensili critici
                            from ml.vita_ottimale import soglia_alert as _sa, soglia_fin_vita as _sfv
                            _records_ctx = _th_load(_cfg_ctx_data)
                            contesto_avvio["utensili_sotto_soglia_alert"] = sum(
                                1 for t in _tools
                                if t.get("life_percent") is not None and
                                t["life_percent"] < _sa((t.get("name") or ""), _records_ctx)
                            )
                            contesto_avvio["utensili_sotto_fin_vita"] = sum(
                                1 for t in _tools
                                if t.get("life_percent") is not None and
                                t["life_percent"] < _sfv((t.get("name") or ""), _records_ctx)
                            )
                            # Mantieni anche le soglie fisse per backward compatibility
                            contesto_avvio["utensili_sotto_20"]  = sum(
                                1 for t in _tools
                                if t.get("life_percent") is not None and t["life_percent"] < 20
                            )
                            contesto_avvio["utensili_sotto_40"]  = sum(
                                1 for t in _tools
                                if t.get("life_percent") is not None and t["life_percent"] < 40
                            )
                            contesto_avvio["vita_media"]         = round(
                                sum(t.get("life_percent") or 0 for t in _tools) / len(_tools), 1
                            ) if _tools else None
                        # Stato pallet adiacenti
                        from api.routers.pallet import _load as _pal_load
                        _pal_state = _pal_load(_cfg_ctx_data)
                        _pallets = _pal_state.get("pallet", [])
                        contesto_avvio["pallet_grezzi"]   = sum(1 for p in _pallets if p.get("stato") == "grezzo")
                        contesto_avvio["pallet_finiti"]   = sum(1 for p in _pallets if p.get("stato") == "finito")
                        contesto_avvio["pallet_vuoti"]    = sum(1 for p in _pallets if p.get("stato") == "vuoto")
                        # Ora e giorno
                        _dt_now = datetime.fromisoformat(now)
                        contesto_avvio["ora_avvio"]       = _dt_now.hour
                        contesto_avvio["giorno_settimana"]= _dt_now.weekday()  # 0=lun, 6=dom
                    except Exception as _e:
                        contesto_avvio["_errore"] = str(_e)

                    sessione = {
                        "id":            sess_id,
                        "data":          now[:10],
                        "progetto":      progetto_nome or "—",
                        "progetto_id":   progetto_id or "",
                        "pallet":        pallet_num,
                        "inizio":        now,
                        "fine":          None,
                        "durata_sec":    None,
                        "programmi":     [],
                        "gap_sec":       0,
                        "utensili":      {},
                        "contesto_avvio": contesto_avvio,  # snapshot ML
                    }
                    data["sessioni"].append(sessione)
                    sc["sessione_id"]  = sess_id
                    sc["inizio_fermo"] = None

                # Avvia nuovo programma
                sc["programma_corrente"]  = programma_attivo
                sc["inizio_programma"]    = now
                sc["utensile_programma"]  = utensile
                sc["t_number_programma"]  = t_number
                sc["in_esecuzione"]       = True
                sc["ultimo_tick"]         = now
                dirty = True

            else:
                # Stesso programma — accumula tempo reale trascorso dall'ultimo tick
                # IMPORTANTE: usa delta temporale reale, NON +5 fisso.
                # Con +5 fisso e N client connessi il valore veniva moltiplicato per N.
                ultimo_tick = sc.get("ultimo_tick")
                if ultimo_tick:
                    try:
                        delta_sec = int((datetime.fromisoformat(now) -
                                        datetime.fromisoformat(ultimo_tick)).total_seconds())
                        delta_sec = max(0, min(delta_sec, 30))  # cap 30s per anomalie
                    except Exception:
                        delta_sec = 0
                else:
                    delta_sec = 0

                if sc.get("sessione_id") and utensile and delta_sec > 0:
                    sess = _find_sess(data, sc["sessione_id"])
                    if sess:
                        sess.setdefault("utensili", {})
                        sess["utensili"][utensile] = sess["utensili"].get(utensile, 0) + delta_sec
                        dirty = True
                # Registra secondi con override ridotto (feed < 90% o mandrino < 90%)
                if sc.get("sessione_id") and delta_sec > 0:
                    ovr_f = override_feed     if override_feed     is not None else 100
                    ovr_m = override_mandrino if override_mandrino is not None else 100
                    if ovr_f < 90 or ovr_m < 90:
                        sess = _find_sess(data, sc["sessione_id"])
                        if sess:
                            sess["sec_override_ridotto"] = sess.get("sec_override_ridotto", 0) + delta_sec
                            min_ovr = min(ovr_f, ovr_m)
                            if "min_override" not in sess or min_ovr < sess["min_override"]:
                                sess["min_override"] = min_ovr
                            dirty = True
                sc["ultimo_tick"] = now

            # ── Rilevamento anomalia ciclo in tempo reale ─────────────────────────
            # Ogni tick confronta elapsed con la media storica del programma.
            # Se elapsed > media + 2σ (e n >= 3 campioni) → flag anomalia_ciclo.
            # Questo viene propagato al frontend nella prossima chiamata sessione-live.
            if sc.get("inizio_programma") and sc.get("programma_corrente"):
                try:
                    elapsed = int((datetime.fromisoformat(now) -
                                   datetime.fromisoformat(sc["inizio_programma"])).total_seconds())
                    fname = sc["programma_corrente"].upper()
                    idx   = data.get("cicli_utensile", {})
                    entry = idx.get(fname)
                    if entry and entry.get("n", 0) >= CICLI_MIN_ANOMALIA:
                        soglia = entry["media"] + CICLI_ANOMALIA_SIGMA * entry["std"]
                        sc["anomalia_ciclo"] = elapsed > soglia
                        sc["anomalia_soglia_sec"] = int(soglia)
                        sc["anomalia_elapsed_sec"] = elapsed
                    else:
                        sc["anomalia_ciclo"] = False
                except Exception:
                    sc["anomalia_ciclo"] = False

    except Exception as _e:
        from utils.logger import get_logger as _get_log
        _get_log("routers.report").warning(f"aggiorna_da_log errore parziale: {_e}")
    finally:
        if dirty:
            _save_log(config, data)

def _find_sess(data: dict, sess_id: str) -> Optional[dict]:
    for s in data.get("sessioni", []):
        if s.get("id") == sess_id:
            return s
    return None

def _chiudi_programma(data: dict, sc: dict, now: str):
    sess = _find_sess(data, sc["sessione_id"])
    if not sess:
        return
    inizio = sc.get("inizio_programma")
    if not inizio:
        return
    try:
        durata = int((datetime.fromisoformat(now) -
                      datetime.fromisoformat(inizio)).total_seconds())
    except Exception:
        durata = None

    info = _parse_nome(sc["programma_corrente"])
    fname = sc["programma_corrente"]

    # ── Fix 2: merge con record precedente se stesso filename (o stesso prefisso) e gap breve
    # Copre due scenari:
    # a) stesso filename esatto (es. pausa/ripresa dello stesso programma)
    # b) stesso prefisso numerico (es. 4297_0006_001 e 4297_0006_801 → stesso lavoro,
    #    diverso suffix perché il Siemens assegna numerazione interna diversa)
    MERGE_GAP_SEC = 900  # 15 minuti
    pgm_prec = sess["programmi"][-1] if sess["programmi"] else None

    def _prefix_norm(fn):
        """Rimuove il suffix numerico Siemens solo se >= 800 (stile _801, _900).
        NON tocca suffix sequenziali normali (001, 002...) che sono programmi distinti.
        Questo evita di mergiare 4350_0221_01_001 con 4350_0221_01_002."""
        base = (fn or "").upper().replace(".MPF", "")
        parts = base.split("_")
        if parts and parts[-1].isdigit() and len(parts[-1]) >= 3 and int(parts[-1]) >= 800:
            return "_".join(parts[:-1])
        return base

    stesso_pgm = (
        pgm_prec is not None
        and pgm_prec.get("fine") is not None
        and durata is not None
        and (
            (pgm_prec.get("filename") or "").upper() == fname.upper()
            or (
                _prefix_norm(pgm_prec.get("filename")) == _prefix_norm(fname)
                and _prefix_norm(fname) != ""
            )
        )
    )
    if stesso_pgm:
        try:
            gap = int((datetime.fromisoformat(inizio) -
                       datetime.fromisoformat(pgm_prec["fine"])).total_seconds())
        except Exception:
            gap = MERGE_GAP_SEC + 1

        if 0 <= gap <= MERGE_GAP_SEC:
            # Merge: somma durate, usa inizio del primo e fine attuale
            dur_prec = pgm_prec.get("durata_sec") or 0
            nuova_durata = dur_prec + durata  # esclude il gap di pausa
            pgm_prec["fine"]       = now
            pgm_prec["durata_sec"] = nuova_durata
            pgm_prec["_merged"]    = pgm_prec.get("_merged", 0) + 1  # traccia quante fusioni
            # Aggiorna indice cicli con la durata corretta (non i frammenti)
            if nuova_durata and 10 <= nuova_durata <= 28800:
                idx = data.setdefault("cicli_utensile", {})
                key_f  = fname.upper()
                idx.setdefault(key_f, {"campioni": [], "n": 0, "media": 0, "std": 0})
                # Rimuovi il campione precedente (era un frammento) e aggiungi quello corretto
                if idx[key_f]["campioni"] and idx[key_f]["campioni"][-1] == dur_prec:
                    idx[key_f]["campioni"][-1] = nuova_durata
                else:
                    _aggiorna_media_incrementale(idx[key_f], nuova_durata)
                idx[key_f]["n"] = len(idx[key_f]["campioni"])
                idx[key_f]["media"] = round(sum(idx[key_f]["campioni"]) / max(1, idx[key_f]["n"]))
            return  # merge completato — non aggiungere nuovo record

    sess["programmi"].append({
        "filename":  fname,
        **info,
        "inizio":    inizio,
        "fine":      now,
        "durata_sec": durata,
        "utensile":  sc.get("utensile_programma"),
        "t_number":  sc.get("t_number_programma"),
    })

    # ── Aggiorna indice cicli per (utensile, filename) ────────────────────────
    # Questo indice permette di rilevare cicli anomali in tempo reale.
    # Soglia: durata valida = 10s … 8h (filtra errori di timing)
    if durata and 10 <= durata <= 28800:
        fname   = sc["programma_corrente"].upper()
        utensile = (sc.get("utensile_programma") or "").upper().strip()
        idx = data.setdefault("cicli_utensile", {})

        # Chiave 1: per filename (indipendente dall'utensile — per ETA programma)
        idx.setdefault(fname, {"campioni": [], "n": 0, "media": 0, "std": 0})
        _aggiorna_media_incrementale(idx[fname], durata)

        # Chiave 2: per coppia (utensile, filename) — per anomalie utensile specifico
        if utensile:
            chiave_ut = f"{utensile}|{fname}"
            idx.setdefault(chiave_ut, {"campioni": [], "n": 0, "media": 0, "std": 0})
            _aggiorna_media_incrementale(idx[chiave_ut], durata)


def _aggiorna_media_incrementale(entry: dict, nuovo_valore: float):
    """
    Mantiene media e deviazione standard incrementale (algoritmo Welford).
    Mantiene gli ultimi 50 campioni per evitare crescita illimitata del file.

    Fix 3: scarta campioni che sono probabilmente frammenti da pausa/interruzione.
    Un campione è sospetto se è < 30% della media storica esistente (con almeno 3 campioni).
    Questo discrimina ripetizioni reali da continuazioni di un ciclo interrotto.
    """
    campioni = entry.get("campioni", [])

    # Scarta frammenti: se abbiamo già ≥3 campioni e la media è affidabile,
    # un valore < 30% della media è quasi certamente un frammento da interruzione
    if len(campioni) >= 3 and entry.get("media", 0) > 0:
        soglia_minima = entry["media"] * 0.30
        if nuovo_valore < soglia_minima:
            return  # scarta — non aggiornare l'indice

    campioni.append(nuovo_valore)
    if len(campioni) > CICLI_FINESTRA:
        campioni = campioni[-CICLI_FINESTRA:]
    entry["campioni"] = campioni
    entry["n"]        = len(campioni)
    entry["media"]    = round(sum(campioni) / len(campioni))
    if len(campioni) > 1:
        media = entry["media"]
        varianza = sum((x - media) ** 2 for x in campioni) / (len(campioni) - 1)
        entry["std"] = round(varianza ** 0.5)
    else:
        entry["std"] = 0

def _chiudi_sessione(data: dict, sc: dict, now: str):
    if not sc.get("sessione_id"):
        return
    _chiudi_programma(data, sc, now)
    sess = _find_sess(data, sc["sessione_id"])
    if not sess:
        return
    # durata_sec = somma programmi chiusi (tempo macchina reale)
    # NON usare (fine - inizio): se la sessione è rimasta aperta nel log
    # mentre la macchina era ferma, il wall-clock include ore di inattività.
    pgms_chiusi = [p for p in sess.get("programmi", []) if p.get("fine") and (p.get("durata_sec") or 0) > 0]
    sess["durata_sec"] = sum(p.get("durata_sec") or 0 for p in pgms_chiusi)
    sess["fine"] = now
    # Calcola gap totale (fermo tra programmi)
    programmi = sess.get("programmi", [])
    gap = 0
    for i in range(1, len(programmi)):
        try:
            prev_fine = programmi[i-1]["fine"]
            curr_inizio = programmi[i]["inizio"]
            if prev_fine and curr_inizio:
                g = int((datetime.fromisoformat(curr_inizio) -
                         datetime.fromisoformat(prev_fine)).total_seconds())
                if g > 0:
                    gap += g
        except Exception:
            pass
    sess["gap_sec"] = gap

# ── Endpoint: dati giornalieri ────────────────────────────────────────────────

@router.get("/fermi/non-classificati")
async def get_fermi_non_classificati():
    """
    Ritorna i fermi globali non ancora classificati dall'operatore.
    Usato dal frontend per mostrare il popup ML al momento giusto.
    Solo fermi >= 10 minuti, completati (hanno fine), non classificati.
    """
    config = carica_configurazione()
    data   = _load_log(config)
    soglia_sec = 600  # 10 minuti

    non_class = [
        f for f in data.get("fermi_globali", [])
        if f.get("fine")                        # fermo completato
        and f.get("durata_sec", 0) >= soglia_sec  # >= 10 minuti
        and not f.get("causa")                  # non ancora classificato
        and not f.get("ignorato")               # non ignorato
    ]
    # Solo gli ultimi 3 non classificati — evita flood popup
    return {"fermi": non_class[-3:]}


@router.post("/fermi/ignora")
async def ignora_fermo(body: dict = Body(...)):
    """Marca un fermo come ignorato (operatore non vuole classificarlo)."""
    config   = carica_configurazione()
    data     = _load_log(config)
    ts_inizio = body.get("ts_inizio")
    if not ts_inizio:
        from fastapi import HTTPException
        raise HTTPException(400, "ts_inizio richiesto")
    for f in data.get("fermi_globali", []):
        if f.get("inizio") == ts_inizio:
            f["ignorato"] = True
            f["ignorato_ts"] = datetime.now().isoformat(timespec="seconds")
            break
    _save_log(config, data)
    return {"ok": True}


@router.post("/fermi/classifica")
async def classifica_fermo(body: dict = Body(...)):
    """
    Classifica la causa di un fermo macchina.
    Salva in lavorazioni_log.json nel campo fermi_classificati.
    causa: setup | allarme | pausa_operatore | altro
    ts_inizio: timestamp ISO inizio fermo
    durata_sec: durata fermo in secondi
    """
    config  = carica_configurazione()
    data    = _load_log(config)
    causa   = (body.get("causa") or "").strip()
    CAUSE_VALIDE = {"setup", "allarme", "pausa_operatore", "altro"}
    if causa not in CAUSE_VALIDE:
        from fastapi import HTTPException
        raise HTTPException(400, f"causa non valida. Valide: {CAUSE_VALIDE}")

    ts_inizio  = body.get("ts_inizio")
    durata_sec = body.get("durata_sec")

    # Salva in fermi_classificati (storico ML)
    fermi = data.setdefault("fermi_classificati", [])
    fermi.append({
        "ts":         datetime.now().isoformat(timespec="seconds"),
        "ts_inizio":  ts_inizio,
        "durata_sec": durata_sec,
        "causa":      causa,
    })

    # Aggiorna anche il record corrispondente in fermi_globali
    # Cerca il fermo globale con ts_inizio più vicino (entro 2 minuti)
    if ts_inizio:
        try:
            ts_ref = datetime.fromisoformat(ts_inizio)
            for fg in data.get("fermi_globali", []):
                if not fg.get("inizio"):
                    continue
                ts_fg = datetime.fromisoformat(fg["inizio"])
                delta = abs((ts_fg - ts_ref).total_seconds())
                if delta <= 120:  # entro 2 minuti
                    fg["causa"] = causa
                    fg["classificato_ts"] = datetime.now().isoformat(timespec="seconds")
                    break
        except Exception:
            pass

    _save_log(config, data)
    return {"ok": True, "causa": causa}


@router.get("/confronto-ieri")
async def get_confronto_ieri():
    """
    Ritorna confronto completati oggi vs ieri alla stessa ora.
    Usato dalla Home per il badge ↑/↓.
    """
    from datetime import timedelta
    config  = carica_configurazione()
    log     = _load_log(config)
    ora_ora = datetime.now()
    oggi    = ora_ora.strftime("%Y-%m-%d")
    ieri    = (ora_ora - timedelta(days=1)).strftime("%Y-%m-%d")
    ora_hm  = ora_ora.strftime("%H:%M")

    def _completati_entro(data_str: str, entro_hm: str) -> int:
        n = 0
        for s in log.get("sessioni", []):
            if s.get("data") != data_str:
                continue
            for p in s.get("programmi", []):
                fine = (p.get("fine") or "")
                if fine and fine[11:16] <= entro_hm:
                    n += 1
        return n

    oggi_n = _completati_entro(oggi, ora_hm)
    ieri_n = _completati_entro(ieri, ora_hm)
    return {
        "oggi": oggi_n,
        "ieri": ieri_n,
        "delta": oggi_n - ieri_n,
        "trend": "su" if oggi_n > ieri_n else "giu" if oggi_n < ieri_n else "pari",
    }


@router.get("/giornaliero")
async def get_report_giornaliero(data: str = Query(default=None)):
    """
    Restituisce report giornaliero.
    data = YYYY-MM-DD (default: oggi)
    """
    config = carica_configurazione()
    log    = _load_log(config)
    target = data or datetime.now().strftime("%Y-%m-%d")
    now_iso = datetime.now().isoformat(timespec="seconds")

    sessioni_giorno = [
        s for s in log.get("sessioni", [])
        if s.get("data") == target
    ]

    def _durata_effettiva(s):
        """Durata reale: usa durata_sec se chiusa, altrimenti calcola live da inizio."""
        if s.get("durata_sec") is not None:
            return s["durata_sec"]
        # Sessione ancora aperta — calcola da inizio a adesso
        inizio = s.get("inizio")
        if not inizio:
            return 0
        try:
            return int((datetime.fromisoformat(now_iso) -
                        datetime.fromisoformat(inizio)).total_seconds())
        except Exception:
            return 0

    def _programmi_effettivi(s):
        """Programmi della sessione, aggiungendo durata live all'ultimo se ancora aperto."""
        pgms = list(s.get("programmi", []))
        sc = log.get("stato_corrente", {})
        # Se questa è la sessione corrente e c'è un programma in corso non chiuso
        if s.get("id") == sc.get("sessione_id") and sc.get("programma_corrente"):
            prog_corrente = sc["programma_corrente"]
            inizio_pgm = sc.get("inizio_programma")
            # Controlla se il programma corrente non è già nell'elenco chiuso
            già_chiuso = any(p.get("filename","").upper().replace(".MPF","") ==
                             prog_corrente.upper().replace(".MPF","") and p.get("fine")
                             for p in pgms)
            if not già_chiuso and inizio_pgm:
                try:
                    dur = int((datetime.fromisoformat(now_iso) -
                               datetime.fromisoformat(inizio_pgm)).total_seconds())
                except Exception:
                    dur = None
                info = _parse_nome(prog_corrente) if prog_corrente else {}
                pgms = pgms + [{
                    "filename":    prog_corrente,
                    **info,
                    "inizio":      inizio_pgm,
                    "fine":        None,
                    "durata_sec":  dur,
                    "utensile":    sc.get("utensile_programma"),
                    "t_number":    sc.get("t_number_programma"),
                    "_live":       True,
                }]
        return pgms

    # Aggregazioni con durata live
    ore_totali  = sum(_durata_effettiva(s) for s in sessioni_giorno)
    gap_sessioni = sum(s.get("gap_sec") or 0 for s in sessioni_giorno)
    n_programmi = sum(len(_programmi_effettivi(s)) for s in sessioni_giorno)

    # Fermo accumulato nello stato_corrente (macchina ferma senza sessione aperta).
    # fermo_sc = tutto il fermo del giorno tracciato tick-per-tick.
    # gap_sessioni = fermi *dentro* le sessioni (gap tra programmi).
    # Il fermo *tra* sessioni (setup, pausa pranzo...) è catturato da fermo_sc
    # ma NON da gap_sessioni. La misura corretta del fermo totale è:
    #   fermo_totale = gap_sessioni + fermo_pre/post_sessione
    # Stima fermo_pre/post: max(0, fermo_sc - gap_sessioni)
    # (fermo_sc include tutto; gap_sessioni è la parte dentro le sessioni)
    sc_oggi = log.get("stato_corrente", {})
    fermo_sc = 0
    if sc_oggi.get("fermo_data") == target:
        fermo_sc = sc_oggi.get("fermo_sec_giornaliero", 0)

    # Per i giorni precedenti (fermo_sc==0) usa fermi_globali come fonte primaria.
    # I fermi_globali sono persistenti e coprono tutti i fermi tra sessioni.
    # Tronca i fermi che attraversano mezzanotte alla sola quota del giorno target.
    if fermo_sc == 0:
        target_start = datetime.fromisoformat(target + "T00:00:00")
        target_end   = datetime.fromisoformat(target + "T23:59:59")
        fermo_sc_acc = 0
        for f in log.get("fermi_globali", []):
            if f.get("ignorato"):
                continue
            # Accetta fermi con data==target OPPURE che si sovrappongono al target
            try:
                f_inizio = datetime.fromisoformat(f["inizio"])
                f_fine   = datetime.fromisoformat(f["fine"]) if f.get("fine") else datetime.fromisoformat(now_iso)
                # Intersezione con il giorno target
                eff_inizio = max(f_inizio, target_start)
                eff_fine   = min(f_fine,   target_end)
                if eff_fine > eff_inizio:
                    fermo_sc_acc += int((eff_fine - eff_inizio).total_seconds())
            except Exception:
                if f.get("data") == target:
                    fermo_sc_acc += f.get("durata_sec", 0)
        fermo_sc = fermo_sc_acc

    fermo_extra = max(0, fermo_sc - gap_sessioni)  # fermo fuori sessioni
    if not sessioni_giorno:
        gap_totale = fermo_sc
    else:
        gap_totale = gap_sessioni + fermo_extra

    # ── OEE — Overall Equipment Effectiveness ─────────────────────────────────
    # Turno 24h = 86400s.
    OEE_TURNO_SEC = ORE_TURNO_SEC
    tempo_turno = max(ore_totali + gap_totale, OEE_TURNO_SEC)

    # Disponibilità: tempo macchina in produzione / tempo turno
    disponibilita = ore_totali / tempo_turno if tempo_turno > 0 else 0

    # Fermi anomali (reset/spegnimento improvviso) vs pianificati (M0/M1)
    n_fermi_anomali   = sum(s.get("n_fermi_anomali", 0)   for s in sessioni_giorno)
    n_fermi_pianif    = sum(s.get("n_fermi_pianificati", 0) for s in sessioni_giorno)

    # Performance: paragona cicli reali con tempi stimati CAM
    # Se non ci sono tempi stimati usa disponibilità come proxy
    sec_teorici = sum(
        pgm.get("durata_sec_teorica") or pgm.get("durata_sec") or 0
        for s in sessioni_giorno
        for pgm in _programmi_effettivi(s)
        if pgm.get("durata_sec")
    )
    sec_reali = sum(
        pgm.get("durata_sec") or 0
        for s in sessioni_giorno
        for pgm in _programmi_effettivi(s)
        if pgm.get("durata_sec")
    )
    performance = min(1.0, sec_teorici / sec_reali) if sec_reali > 0 and sec_teorici > 0 else disponibilita

    # Qualità: proxy per CNC stampi (rarissimi scarti) = 98%
    qualita = OEE_QUALITA_DEFAULT

    oee = round(disponibilita * performance * qualita * 100, 1)

    # Override ridotto — tempo totale con feed/mandrino < 90%
    sec_ovr_ridotto = sum(s.get("sec_override_ridotto") or 0 for s in sessioni_giorno)
    min_ovr_giorno  = min(
        (s["min_override"] for s in sessioni_giorno if "min_override" in s),
        default=None
    )

    # ── Perdite TPM classificate ───────────────────────────────────────────────
    # Modello 6 grandi perdite adattato a CNC monomacchina:
    # 1. Guasti/reset anomali     → n_fermi_anomali × durata_media_fermo
    # 2. Setup/cambio pallet      → gap tra fine sessione e inizio successiva
    # 3. Microfermi tra programmi → gap_sec dentro sessione (cambio utensile, misura)
    # 4. Velocità ridotta         → sec_override_ridotto (feed/mandrino < 90%)
    # 5. Produzione netta         → ore_totali - velocità ridotta
    # 6. Tempo libero turno       → ore_turno - tutto il resto

    # Gap setup: tempo tra fine di una sessione e inizio della successiva (stesso giorno)
    sec_setup = 0
    sessioni_sorted = sorted(
        [s for s in sessioni_giorno if s.get("inizio") and s.get("fine")],
        key=lambda s: s["inizio"]
    )
    for i in range(1, len(sessioni_sorted)):
        prev_fine   = sessioni_sorted[i-1].get("fine")
        curr_inizio = sessioni_sorted[i].get("inizio")
        if prev_fine and curr_inizio:
            try:
                g = int((datetime.fromisoformat(curr_inizio) -
                         datetime.fromisoformat(prev_fine)).total_seconds())
                if 60 <= g <= 7200:   # tra 1 min e 2h — filtra valori anomali
                    sec_setup += g
            except Exception:
                pass

    # Microfermi: gap tra programmi *dentro* ogni sessione.
    # NON usa gap_totale perché ora include fermo_extra (fermo fuori sessioni)
    # che appartiene alla categoria "setup/libero", non ai microfermi.
    sec_microfermi = gap_sessioni

    # Stima durata media fermo anomalo (se disponibile)
    sec_fermi_anomali = 0
    for s in sessioni_giorno:
        fermi = s.get("fermi", [])
        n_anom = s.get("n_fermi_anomali", 0)
        if n_anom > 0 and fermi:
            anom = [f for f in fermi if f.get("tipo") == "reset"]
            sec_fermi_anomali += len(anom) * 300  # stima 5 min per fermo anomalo
        elif n_anom > 0:
            sec_fermi_anomali += n_anom * 300

    sec_libero = max(0, tempo_turno - ore_totali - sec_setup - sec_fermi_anomali - fermo_extra)
    sec_produzione_netta = max(0, ore_totali - sec_ovr_ridotto)

    perdite_tpm = {
        "produzione_netta_sec":  sec_produzione_netta,
        "velocita_ridotta_sec":  sec_ovr_ridotto,
        "microfermi_sec":        sec_microfermi,
        "setup_sec":             sec_setup,
        "guasti_sec":            sec_fermi_anomali,
        "fermo_extra_sec":       fermo_extra,   # fermo pre/post sessione (attesa, pranzo...)
        "libero_sec":            sec_libero,
        "n_setup":               max(0, len(sessioni_sorted) - 1),
        "n_guasti":              n_fermi_anomali,
        "media_setup_sec":       round(sec_setup / max(1, len(sessioni_sorted) - 1))
                                 if len(sessioni_sorted) > 1 else 0,
    }

    # Utensili aggregati (usa accumulo tick — già corretto anche live)
    utensili_agg = {}
    for s in sessioni_giorno:
        for ut, sec in (s.get("utensili") or {}).items():
            utensili_agg[ut] = utensili_agg.get(ut, 0) + sec

    # Progetti aggregati con durata live
    progetti_agg = {}
    for s in sessioni_giorno:
        prog = s.get("progetto") or "—"
        if prog not in progetti_agg:
            progetti_agg[prog] = {"durata_sec": 0, "n_programmi": 0, "pallet": s.get("pallet")}
        progetti_agg[prog]["durata_sec"] += _durata_effettiva(s)
        progetti_agg[prog]["n_programmi"] += len(_programmi_effettivi(s))

    # Sessioni con programmi arricchiti (live) e durata calcolata
    sessioni_output = []
    for s in sessioni_giorno:
        s_out = dict(s)
        s_out["programmi"]  = _programmi_effettivi(s)
        s_out["durata_sec"] = _durata_effettiva(s)
        sessioni_output.append(s_out)

    return {
        "data":             target,
        "ore_lavorate":     _durata_str(ore_totali),
        "ore_lavorate_sec": ore_totali,
        "tempo_fermo_sec":  gap_totale,
        "tempo_fermo":      _durata_str(gap_totale),
        "n_sessioni":       len(sessioni_giorno),
        "n_programmi":      n_programmi,
        "efficienza_pct":   round(ore_totali / (ore_totali + gap_totale) * 100, 1)
                            if (ore_totali + gap_totale) > 0 else 0,
        # OEE
        "oee": {
            "valore":              oee,
            "disponibilita":       round(disponibilita * 100, 1),
            "performance":         round(performance   * 100, 1),
            "qualita":             round(qualita       * 100, 1),
            "ore_turno_sec":       tempo_turno,
            "n_fermi_anomali":     n_fermi_anomali,
            "n_fermi_pianificati": n_fermi_pianif,
        },
        # Override ridotto
        "override_ridotto": {
            "sec_totale":    sec_ovr_ridotto,
            "durata":        _durata_str(sec_ovr_ridotto),
            "pct_tempo":     round(sec_ovr_ridotto / ore_totali * 100, 1) if ore_totali > 0 else 0,
            "min_valore":    min_ovr_giorno,
        },
        # Perdite TPM classificate
        "perdite_tpm":      perdite_tpm,
        "progetti":         progetti_agg,
        "utensili":         {k: {"sec": v, "ore": _durata_str(v)}
                              for k, v in sorted(utensili_agg.items(), key=lambda x: -x[1])},
        "sessioni":         sessioni_output,
        "fermi_globali":    [f for f in log.get("fermi_globali", []) if f.get("data") == target],
    }

@router.get("/storico")
async def get_storico(giorni: int = Query(default=7, ge=1, le=365)):
    """Ultimi N giorni — per grafici trend. Max 365."""
    config  = carica_configurazione()
    log     = _load_log(config)
    sc_oggi = log.get("stato_corrente", {})
    now_iso = datetime.now().isoformat(timespec="seconds")
    today   = datetime.now().strftime("%Y-%m-%d")
    result  = []
    for i in range(giorni - 1, -1, -1):
        d    = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        sess = [s for s in log.get("sessioni", []) if s.get("data") == d]
        def _dur(s):
            if s.get("durata_sec") is not None:
                return s["durata_sec"]
            if d == today:
                inizio = s.get("inizio")
                if inizio:
                    try:
                        return int((datetime.fromisoformat(now_iso) -
                                    datetime.fromisoformat(inizio)).total_seconds())
                    except Exception:
                        pass
            return 0
        ore = sum(_dur(s) for s in sess)
        gap_sess = sum(s.get("gap_sec") or 0 for s in sess)

        # Per oggi: usa fermo_sc tick-per-tick
        # Per giorni precedenti: usa fermi_globali (persistenti)
        fermo_sc = 0
        if d == today and sc_oggi.get("fermo_data") == today:
            fermo_sc = sc_oggi.get("fermo_sec_giornaliero", 0)
        if fermo_sc == 0:
            fermi_del_giorno = [
                f for f in log.get("fermi_globali", [])
                if f.get("data") == d and not f.get("ignorato")
            ]
            fermo_sc = sum(f.get("durata_sec", 0) for f in fermi_del_giorno)
        # Somma corretta: gap dentro sessioni + fermo fuori sessioni
        fermo_extra = max(0, fermo_sc - gap_sess)
        gap = gap_sess + fermo_extra if sess else fermo_sc

        ORE_TURNO = 86400
        tempo_turno   = max(ore + gap, ORE_TURNO)
        disponibilita = ore / tempo_turno if tempo_turno > 0 else 0
        oee_valore    = round(disponibilita * 1.0 * 0.98 * 100, 1)
        sec_ovr       = sum(s.get("sec_override_ridotto") or 0 for s in sess)
        n_fermi_anom  = sum(s.get("n_fermi_anomali", 0) for s in sess)
        n_fermi_pian  = sum(s.get("n_fermi_pianificati", 0) for s in sess)

        result.append({
            "data":               d,
            "ore_lavorate_sec":   ore,
            "ore_lavorate":       _durata_str(ore),
            "tempo_fermo_sec":    gap,
            "tempo_fermo":        _durata_str(gap),
            "n_programmi":        sum(len(s.get("programmi", [])) for s in sess),
            "n_sessioni":         len(sess),
            "efficienza_pct":     round(ore / (ore + gap) * 100, 1) if (ore + gap) > 0 else 0,
            "oee":                {"valore": oee_valore, "disponibilita": round(disponibilita*100,1)} if ore > 0 else None,
            "sec_override_ridotto": sec_ovr,
            "n_fermi_anomali":    n_fermi_anom,
            "n_fermi_pianificati": n_fermi_pian,
            # flag: fermo_sc usato per oggi (utile per tooltip nel frontend)
            "fermo_da_sc":        fermo_sc > gap_sess if d == today else False,
        })
    return result

# ── Sessione live ─────────────────────────────────────────────────────────────

@router.get("/sessione-live")
async def get_sessione_live():
    """
    Restituisce i dati della sessione di lavorazione attualmente in corso.
    Chiamato dalla Home ogni 10s per mostrare il timer del pallet attivo.

    Risposta:
    {
      "attiva": true,
      "inizio_sessione": "2026-03-30T08:12:00",
      "durata_sec": 7543,
      "durata_str": "02:05:43",
      "inizio_programma": "2026-03-30T10:16:46",
      "durata_programma_sec": 312,
      "programma_corrente": "4297_005_01_14.MPF",
      "utensile": "FS16R2L80F85E6",
      "n_programmi_sessione": 14
    }
    """
    config  = carica_configurazione()
    data    = _load_log(config)
    sc      = data.get("stato_corrente", {})
    now_iso = datetime.now().isoformat(timespec="seconds")

    # Tempo fermo giornaliero — disponibile sempre, anche senza sessione
    today      = now_iso[:10]
    fermo_oggi = sc.get("fermo_sec_giornaliero", 0) if sc.get("fermo_data") == today else 0

    # Durante grace period (in_pausa): sessione aperta ma non in esecuzione
    in_pausa = sc.get("in_pausa", False)
    if not sc.get("in_esecuzione") and not in_pausa:
        return {"attiva": False, "fermo_sec_giornaliero": fermo_oggi}
    if not sc.get("sessione_id"):
        return {"attiva": False, "fermo_sec_giornaliero": fermo_oggi}

    sess = _find_sess(data, sc["sessione_id"])
    if not sess:
        return {"attiva": False, "fermo_sec_giornaliero": fermo_oggi}

    # Durata sessione = somma programmi chiusi + programma in corso live
    # NON usare (now - inizio_sessione): se la sessione rimane aperta nel log
    # mentre la macchina è ferma, il wall-clock conta ore di inattività.
    pgms = sess.get("programmi", [])
    durata_pgm_chiusi = sum(p.get("durata_sec") or 0 for p in pgms if p.get("fine"))
    inizio_sess = sess.get("inizio")  # tenuto per compatibilità frontend

    # Programma corrente in esecuzione — aggiunge elapsed live
    inizio_pgm = sc.get("inizio_programma")
    durata_pgm = 0
    if inizio_pgm and sc.get("in_esecuzione"):
        try:
            durata_pgm = int((datetime.fromisoformat(now_iso) -
                              datetime.fromisoformat(inizio_pgm)).total_seconds())
            durata_pgm = max(0, durata_pgm)
        except Exception:
            pass

    durata_sess = durata_pgm_chiusi + durata_pgm

    n_pgm = len(sess.get("programmi", []))  # già chiusi + quello corrente
    if sc.get("programma_corrente"):
        n_pgm += 1  # aggiungi quello in corso

    # Statistiche ciclo storico per il programma corrente
    ciclo_stats = None
    fname_curr = (sc.get("programma_corrente") or "").upper()
    if fname_curr:
        idx   = data.get("cicli_utensile", {})
        entry = idx.get(fname_curr)
        if entry and entry.get("n", 0) >= 2:
            ciclo_stats = {
                "media_sec": entry["media"],
                "std_sec":   entry["std"],
                "n":         entry["n"],
            }

    return {
        "attiva":                True,
        "in_pausa":              in_pausa,
        "fermo_sec_giornaliero": fermo_oggi,
        "inizio_sessione":       inizio_sess,
        "durata_sec":            durata_sess,
        "durata_str":            _durata_str(durata_sess),
        "inizio_programma":      inizio_pgm,
        "durata_programma_sec":  durata_pgm,
        "durata_programma_str":  _durata_str(durata_pgm),
        "programma_corrente":    sc.get("programma_corrente"),
        "utensile":              sc.get("utensile_programma"),
        "t_number":              sc.get("t_number_programma"),
        "n_programmi_sessione":  n_pgm,
        "progetto":              sess.get("progetto"),
        "pallet":                sess.get("pallet"),
        "anomalia_ciclo":        sc.get("anomalia_ciclo", False),
        "anomalia_soglia_sec":   sc.get("anomalia_soglia_sec"),
        "anomalia_elapsed_sec":  sc.get("anomalia_elapsed_sec"),
        "ciclo_stats":           ciclo_stats,
    }


@router.post("/reset-sessione")
async def reset_sessione():
    """
    Chiude la sessione corrente e azzera stato_corrente.
    Usato quando la sessione è orfana (pallet/progetto sbagliato).
    Il prossimo tick di aggiorna-stati-da-log creerà una nuova sessione corretta.
    """
    config = carica_configurazione()
    data   = _load_log(config)
    sc     = data.get("stato_corrente", {})
    now    = datetime.now().isoformat(timespec="seconds")

    # Chiudi sessione orfana se presente
    if sc.get("sessione_id"):
        sess = _find_sess(data, sc["sessione_id"])
        if sess and not sess.get("fine"):
            # Calcola durata da programmi chiusi
            pgms_chiusi = [p for p in sess.get("programmi", [])
                           if p.get("fine") and (p.get("durata_sec") or 0) > 0]
            sess["durata_sec"] = sum(p["durata_sec"] for p in pgms_chiusi)
            sess["fine"] = now

    # Azzera stato corrente — il prossimo tick ricostruisce tutto
    data["stato_corrente"] = {
        "fermo_sec_giornaliero": sc.get("fermo_sec_giornaliero", 0),
        "fermo_data": sc.get("fermo_data", ""),
    }
    _save_log(config, data)
    return {"ok": True, "msg": "Sessione resettata — nuovo ciclo al prossimo tick"}


@router.post("/correggi-ore-utensili")
async def correggi_ore_utensili():
    """
    Correzione retroattiva ore utensile gonfiate dalla race condition multi-client.
    
    Causa: N client chiamavano aggiorna-stati-da-log contemporaneamente → +5s×N per tick.
    Fix: per ogni sessione gonfiata, scala le ore utensile al valore corretto
         usando il rapporto durata_programmi / ore_utensili_totali.
    
    Idempotente — applicare più volte non cambia il risultato.
    """
    config = carica_configurazione()
    data   = _load_log(config)
    
    corrette = 0
    dettaglio = []
    
    for s in data.get("sessioni", []):
        pgms = s.get("programmi", [])
        durata_pgms = sum(p.get("durata_sec", 0) for p in pgms)
        utensili = s.get("utensili", {})
        if not utensili or not isinstance(utensili, dict):
            continue
        
        ut_tot = sum(v for v in utensili.values() if isinstance(v, (int, float)))
        
        # Soglia: correggi solo se gonfiato >40% oltre la durata programmi
        if durata_pgms <= 0 or ut_tot <= durata_pgms * 1.4:
            continue
        
        # Già corretto in precedenza? Controlla il flag
        if s.get("utensili_corretti"):
            continue
        
        factor = durata_pgms / ut_tot
        new_utensili = {}
        for alias, sec in utensili.items():
            if isinstance(sec, (int, float)):
                new_utensili[alias] = max(1, int(sec * factor))
            else:
                new_utensili[alias] = sec
        
        s["utensili"] = new_utensili
        s["utensili_corretti"] = True  # flag per idempotenza
        corrette += 1
        dettaglio.append({
            "sessione_id": s["id"],
            "progetto":    s.get("progetto"),
            "ut_prima":    ut_tot,
            "ut_dopo":     sum(v for v in new_utensili.values() if isinstance(v,(int,float))),
            "factor":      round(factor, 3),
        })
    
    if corrette > 0:
        _save_log(config, data)
    
    return {
        "ok":       True,
        "corrette": corrette,
        "dettaglio": dettaglio,
        "msg": f"{corrette} sessioni corrette" if corrette else "Nessuna sessione da correggere",
    }


@router.get("/export-excel")
async def export_excel(data: str = Query(default=None)):
    """Genera report Excel giornaliero."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    config = carica_configurazione()
    target = data or datetime.now().strftime("%Y-%m-%d")
    rpt    = (await get_report_giornaliero(target))

    wb = openpyxl.Workbook()

    # ── Colori ────────────────────────────────────────────────────────────────
    BLU_HEADER = "1D5FAD"
    BLU_LIGHT  = "DBEAFE"
    GRIGIO     = "F1F5F9"
    VERDE      = "16A34A"
    ARANCIONE  = "D97706"

    def hdr(ws, row, col, val, bold=True, bg=BLU_HEADER, fg="FFFFFF", size=11):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=size, name="Arial")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def cell(ws, row, col, val, bold=False, align="left", fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, name="Arial", size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        if fmt: c.number_format = fmt
        return c

    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ════════════════════════════════════════════════════════════════════════
    # Foglio 1 — Riepilogo
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Riepilogo"
    ws1.sheet_view.showGridLines = False

    # Titolo
    ws1.merge_cells("A1:F1")
    t = ws1.cell(row=1, column=1,
                 value=f"REPORT LAVORAZIONE — {target}")
    t.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    t.fill = PatternFill("solid", start_color=BLU_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # KPI box
    kpi = [
        ("Ore lavorate",    rpt["ore_lavorate"]),
        ("Tempo fermo",     rpt["tempo_fermo"]),
        ("Efficienza",      f"{rpt['efficienza_pct']}%"),
        ("N° programmi",    rpt["n_programmi"]),
        ("N° sessioni",     rpt["n_sessioni"]),
    ]
    ws1.row_dimensions[3].height = 15
    for i, (label, val) in enumerate(kpi, 1):
        c = ws1.cell(row=3, column=i, value=label)
        c.font = Font(bold=True, size=9, color="64748B", name="Arial")
        c.alignment = Alignment(horizontal="center")
        c2 = ws1.cell(row=4, column=i, value=val)
        c2.font = Font(bold=True, size=14, name="Arial",
                       color=VERDE if i == 3 else "0D2D5E")
        c2.alignment = Alignment(horizontal="center")
        c2.fill = PatternFill("solid", start_color=BLU_LIGHT)
        c2.border = thin
        ws1.row_dimensions[4].height = 30

    # Sezione Progetti
    r = 7
    ws1.merge_cells(f"A{r}:F{r}")
    hdr(ws1, r, 1, "PROGETTI LAVORATI", bg=BLU_HEADER)
    r += 1
    for col, (lbl, w) in enumerate([
        ("Progetto", 20), ("Pallet", 10), ("Ore lavorate", 16),
        ("N° programmi", 15), ("—", 10), ("—", 10)
    ], 1):
        hdr(ws1, r, col, lbl, bg="334155", size=10)
        ws1.column_dimensions[get_column_letter(col)].width = w
    r += 1
    for prog, info in rpt["progetti"].items():
        fill = PatternFill("solid", start_color=GRIGIO if r % 2 == 0 else "FFFFFF")
        cell(ws1, r, 1, prog, bold=True).fill = fill
        cell(ws1, r, 2, f"P{info['pallet']}" if info.get("pallet") else "—", align="center").fill = fill
        cell(ws1, r, 3, _durata_str(info["durata_sec"]), align="center").fill = fill
        cell(ws1, r, 4, info["n_programmi"], align="center").fill = fill
        for col in [1,2,3,4]:
            ws1.cell(row=r, column=col).border = thin
        r += 1

    # Sezione Utensili
    r += 2
    ws1.merge_cells(f"A{r}:D{r}")
    hdr(ws1, r, 1, "UTILIZZO UTENSILI", bg=BLU_HEADER)
    r += 1
    for col, lbl in enumerate(["Alias utensile", "Ore utilizzo", "% sul totale", "—"], 1):
        hdr(ws1, r, col, lbl, bg="334155", size=10)
    r += 1
    ore_tot = sum(v["sec"] for v in rpt["utensili"].values()) or 1
    for alias, info in rpt["utensili"].items():
        pct = round(info["sec"] / ore_tot * 100, 1)
        fill = PatternFill("solid", start_color=GRIGIO if r % 2 == 0 else "FFFFFF")
        cell(ws1, r, 1, alias, bold=True).fill = fill
        cell(ws1, r, 2, info["ore"], align="center").fill = fill
        cell(ws1, r, 3, f"{pct}%", align="center").fill = fill
        for col in [1,2,3]:
            ws1.cell(row=r, column=col).border = thin
        r += 1

    # ════════════════════════════════════════════════════════════════════════
    # Foglio 2 — Dettaglio programmi
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Dettaglio Programmi")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:I1")
    t2 = ws2.cell(row=1, column=1, value=f"DETTAGLIO PROGRAMMI — {target}")
    t2.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t2.fill = PatternFill("solid", start_color=BLU_HEADER)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    cols2 = [
        ("Progetto", 18), ("Pallet", 8), ("Commessa", 10),
        ("Posizione", 10), ("Fase", 8), ("N°", 8),
        ("Inizio", 18), ("Fine", 18), ("Durata", 12),
        ("Utensile", 20),
    ]
    for i, (lbl, w) in enumerate(cols2, 1):
        hdr(ws2, 2, i, lbl, bg="334155", size=10)
        ws2.column_dimensions[get_column_letter(i)].width = w

    r2 = 3
    for sess in rpt["sessioni"]:
        for pgm in sess.get("programmi", []):
            fill = PatternFill("solid", start_color=GRIGIO if r2 % 2 == 0 else "FFFFFF")
            vals = [
                sess.get("progetto"), f"P{sess.get('pallet','')}" if sess.get("pallet") else "—",
                pgm.get("commessa"), pgm.get("posizione"),
                pgm.get("fase"), pgm.get("seq"),
                pgm.get("inizio","")[:16].replace("T"," "),
                pgm.get("fine","")[:16].replace("T"," ") if pgm.get("fine") else "—",
                _durata_str(pgm.get("durata_sec")),
                pgm.get("utensile") or "—",
            ]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=r2, column=col, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = fill
                c.border = thin
                c.alignment = Alignment(horizontal="center" if col > 2 else "left",
                                        vertical="center")
            r2 += 1

    # ════════════════════════════════════════════════════════════════════════
    # Foglio 3 — Tempi fermo
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Fermi Macchina")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:E1")
    t3 = ws3.cell(row=1, column=1, value=f"FERMI MACCHINA — {target}")
    t3.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t3.fill = PatternFill("solid", start_color=ARANCIONE)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    cols3 = [("Inizio",18),("Fine",18),("Durata",14),("Causa",18),("Note",20)]
    for i,(lbl,w) in enumerate(cols3,1):
        hdr(ws3, 2, i, lbl, bg="92400E", size=10)
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Usa fermi_globali (più accurati) — include fermi tra sessioni
    fermi_src = rpt.get("fermi_globali") or []

    # Fallback: gap tra programmi se fermi_globali vuoti
    if not fermi_src:
        for sess in rpt["sessioni"]:
            pgms = sess.get("programmi",[])
            for i in range(1, len(pgms)):
                prev = pgms[i-1]; curr = pgms[i]
                if not (prev.get("fine") and curr.get("inizio")): continue
                try:
                    gap = int((datetime.fromisoformat(curr["inizio"]) -
                               datetime.fromisoformat(prev["fine"])).total_seconds())
                except: continue
                if gap < 10: continue
                fermi_src.append({
                    "inizio": prev.get("fine"),
                    "fine":   curr.get("inizio"),
                    "durata_sec": gap,
                    "causa": None,
                })

    fermi_src = sorted(fermi_src, key=lambda f: f.get("inizio") or "")

    r3 = 3
    for f in fermi_src:
        dur = f.get("durata_sec", 0)
        if dur < 10: continue
        inizio_str = (f.get("inizio") or "")[:16].replace("T"," ")
        fine_str   = (f.get("fine")   or "")[:16].replace("T"," ")
        causa      = (f.get("causa") or "—")
        fill = PatternFill("solid", start_color="FEF3C7" if r3%2==0 else "FFFBEB")
        # Colore causa
        causa_color = {
            "allarme":         "DC2626",
            "setup":           "1D5FAD",
            "pausa_operatore": "16A34A",
            "altro":           "64748B",
        }.get(causa, "374151")
        vals = [inizio_str, fine_str, _durata_str(dur), causa.replace("_"," "), ""]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=r3, column=col, value=val)
            c.font = Font(name="Arial", size=10,
                         bold=(col==4), color=causa_color if col==4 else "000000")
            c.fill = fill; c.border = thin
            c.alignment = Alignment(horizontal="center", vertical="center")
        r3 += 1

    if r3 == 3:
        c = ws3.cell(row=3, column=1, value="Nessun fermo registrato per questa giornata")
        c.font = Font(name="Arial", size=10, italic=True, color="94A3B8")
        ws3.merge_cells("A3:E3")

    # Salva
    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    out = Path(tmp.name)
    wb.save(str(out))
    return out

@router.get("/export-excel-download")
async def export_excel_download(data: str = Query(default=None)):
    from fastapi.responses import StreamingResponse
    import io
    target = data or datetime.now().strftime("%Y-%m-%d")
    out    = await export_excel(data)
    fname  = f"report_lavorazione_{target}.xlsx"
    # Leggi il file in memoria e restituisci come stream — funziona su qualsiasi OS
    with open(str(out), "rb") as f:
        content_bytes = f.read()
    # Pulizia file temp
    try:
        out.unlink()
    except Exception:
        pass
    return StreamingResponse(
        io.BytesIO(content_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ── Tempi ciclo reali per ETA ─────────────────────────────────────────────────

@router.get("/tempi-ciclo")
async def get_tempi_ciclo():
    """
    Analizza lavorazioni_log.json e calcola statistiche del ciclo reale
    per ogni filename MPF: media, deviazione standard, n campioni.

    Usato per:
    - ETA programma corrente nella Home (media * n_programmi_rimanenti)
    - Aggiornamento automatico tempoStimato nei progetti
    - Rilevamento cicli anomali (durata > media + 2σ)

    Risposta:
    {
      "cicli": {
        "4297_005_01_18.MPF": {
          "media_sec":  342,
          "std_sec":    28,
          "min_sec":    310,
          "max_sec":    398,
          "n":          7,
          "ultimo":     "2026-03-30T22:04:00"
        },
        ...
      },
      "n_programmi_con_dati": 45,
      "sessioni_analizzate":  12
    }
    """
    config = carica_configurazione()
    log    = _load_log(config)
    sessioni = log.get("sessioni", [])

    # Accumula durate per filename
    from collections import defaultdict
    import statistics

    dati: dict[str, list] = defaultdict(list)
    ultimi: dict[str, str] = {}

    for sess in sessioni:
        for pgm in sess.get("programmi", []):
            fname = pgm.get("filename", "").strip().upper()
            dur   = pgm.get("durata_sec")
            fine  = pgm.get("fine")
            if not fname or dur is None or dur <= 0:
                continue
            # Filtra durate anomale (> 8h = quasi certamente un bug di timing)
            if dur > TEMPI_CICLO_MAX_DURATA:
                continue
            dati[fname].append(dur)
            if fine and (fname not in ultimi or fine > ultimi[fname]):
                ultimi[fname] = fine

    cicli = {}
    for fname, durate in dati.items():
        if len(durate) < 1:
            continue
        media = sum(durate) / len(durate)
        std   = statistics.stdev(durate) if len(durate) > 1 else 0
        cicli[fname] = {
            "media_sec": round(media),
            "std_sec":   round(std),
            "min_sec":   min(durate),
            "max_sec":   max(durate),
            "n":         len(durate),
            "ultimo":    ultimi.get(fname),
        }

    return {
        "cicli":                  cicli,
        "n_programmi_con_dati":   len(cicli),
        "sessioni_analizzate":    len(sessioni),
    }


@router.get("/cicli-utensile")
async def get_cicli_utensile(utensile: str = None):
    """
    Ritorna l'indice cicli_utensile dal log — dati live aggiornati ogni completamento.
    Differisce da /tempi-ciclo che ricalcola da zero sulle sessioni storiche:
    questo usa l'indice incrementale (finestra scorrevole 50 campioni).

    Opzionale: ?utensile=FS16R2L85 filtra per utensile specifico.

    Usato per:
    - Pagina diagnostica utensili (tempi ciclo per alias)
    - Rilevamento degradazione utensile (ciclo si allunga nel tempo)
    """
    config = carica_configurazione()
    data   = _load_log(config)
    idx    = data.get("cicli_utensile", {})

    if utensile:
        # Filtra per utensile specifico: chiavi del tipo "ALIAS|FILENAME.MPF"
        filtro = utensile.upper().strip()
        idx = {k: v for k, v in idx.items() if k.startswith(filtro + "|")}

    # Separa indici per filename puro da indici per coppia utensile|filename
    per_file   = {k: v for k, v in idx.items() if "|" not in k}
    per_ut_pgm = {k: v for k, v in idx.items() if "|" in k}

    # Estrai utensili unici con statistiche aggregate + slope degradazione
    utensili_stats = {}
    for chiave, entry in per_ut_pgm.items():
        alias = chiave.split("|")[0]
        if alias not in utensili_stats:
            utensili_stats[alias] = {"n_programmi": 0, "n_cicli": 0, "programmi": [], "_slope_sum": 0.0, "_slope_n": 0}
        us = utensili_stats[alias]
        us["n_programmi"] += 1
        us["n_cicli"]     += entry.get("n", 0)
        campioni = entry.get("campioni", [])
        slope = 0.0
        if len(campioni) >= 3:
            n = len(campioni)
            media_c = sum(campioni) / n
            x_mean = (n - 1) / 2
            num = sum((i - x_mean) * (campioni[i] - media_c) for i in range(n))
            den = sum((x - x_mean)**2 for x in range(n))
            slope = round(num / den, 2) if den > 0 else 0.0
        us["_slope_sum"] += slope * entry.get("n", 1)
        us["_slope_n"]   += entry.get("n", 1)
        us["programmi"].append({
            "filename":  chiave.split("|", 1)[1] if "|" in chiave else chiave,
            "media_sec": entry.get("media", 0),
            "std_sec":   entry.get("std", 0),
            "n":         entry.get("n", 0),
            "slope":     round(slope, 2),
        })

    for alias, us in utensili_stats.items():
        us["slope_medio"] = round(us["_slope_sum"] / max(1, us["_slope_n"]), 2)
        us["degrado"]     = us["slope_medio"] > 2.0
        del us["_slope_sum"], us["_slope_n"]

    return {
        "per_file":       per_file,
        "per_utensile":   utensili_stats,
        "n_file_tracciati": len(per_file),
        "n_utensili":     len(utensili_stats),
    }


@router.get("/cicli-dettaglio")
async def get_cicli_dettaglio(filename: str = None):
    """
    Restituisce i campioni grezzi di ciclo per un programma specifico.
    Usato dalla tab SPC del Report per run chart e calcolo indici statistici.

    ?filename=4297_005_01_18.MPF → campioni, media, std, CV%, p95, trend slope
    Senza parametro → lista programmi con >=5 campioni (per selettore UI)
    """
    config = carica_configurazione()
    data   = _load_log(config)
    idx    = data.get("cicli_utensile", {})

    def _statistiche(campioni: list) -> dict:
        if not campioni:
            return {}
        n = len(campioni)
        media = sum(campioni) / n
        std   = (sum((x - media)**2 for x in campioni) / max(1, n-1)) ** 0.5
        cv    = round(std / media * 100, 1) if media > 0 else 0
        sorted_c = sorted(campioni)
        p95   = sorted_c[int(n * 0.95)] if n >= 5 else sorted_c[-1]
        mediana = sorted_c[n // 2]
        # Slope lineare (regressione semplice) per rilevare trend degradazione
        if n >= 3:
            xs = list(range(n))
            x_mean = (n - 1) / 2
            slope_num = sum((xs[i] - x_mean) * (campioni[i] - media) for i in range(n))
            slope_den = sum((x - x_mean)**2 for x in xs)
            slope = round(slope_num / slope_den, 2) if slope_den > 0 else 0
        else:
            slope = 0
        return {
            "n":        n,
            "media":    round(media),
            "std":      round(std),
            "cv_pct":   cv,
            "p95":      p95,
            "mediana":  mediana,
            "min":      sorted_c[0],
            "max":      sorted_c[-1],
            "slope":    slope,            # sec per ciclo — positivo = degrado
            "stabile":  cv < 10,          # CV < 10% = processo stabile
        }

    if filename:
        fname = filename.upper().strip()
        entry = idx.get(fname)
        if not entry:
            return {"filename": filename, "campioni": [], "stats": {}}
        campioni = entry.get("campioni", [])
        stats = _statistiche(campioni)
        # Ultime 20 per il run chart (più recenti)
        return {
            "filename":  filename,
            "campioni":  campioni[-50:],   # tutti (max 50 per finestra scorrevole)
            "run_chart": campioni[-20:],   # ultimi 20 per visualizzazione
            "stats":     stats,
        }
    else:
        # Lista programmi con sufficienti campioni per analisi SPC
        programmi = []
        for fname, entry in idx.items():
            if "|" in fname:
                continue  # skip coppie utensile|filename
            campioni = entry.get("campioni", [])
            if len(campioni) < 5:
                continue
            stats = _statistiche(campioni)
            programmi.append({
                "filename":  fname,
                "n":         stats["n"],
                "media_sec": stats["media"],
                "cv_pct":    stats["cv_pct"],
                "stabile":   stats["stabile"],
                "slope":     stats["slope"],
            })
        # Ordina per CV% decrescente (i più instabili prima)
        programmi.sort(key=lambda p: -p["cv_pct"])
        return {"programmi": programmi, "n_totale": len(programmi)}


@router.get("/ore-progetto")
async def get_ore_progetto(progetto: str = None, project_id: str = None):
    """
    Ritorna le ore totali accumulate per progetto su TUTTO lo storico.
    Cerca per nome esatto (progetto=) o per project_id= salvato nel log.
    Senza parametri → tutti i progetti.
    """
    config = carica_configurazione()
    data   = _load_log(config)

    # Aggrega da tutte le sessioni (log corrente)
    agg: dict = {}

    def _aggiungi(sessione: dict):
        prog = (sessione.get("progetto") or "").strip()
        if not prog or prog == "—":
            return
        # Salta sessioni aperte (fine=None o "aperta"):
        # il frontend le aggiunge live tramite durataSessioneLive.
        # Se le includiamo qui si conterebbero due volte.
        fine = sessione.get("fine") or ""
        if not fine or fine == "aperta":
            return
        # Salta sessioni durata zero — record spazzatura
        dur_check = sessione.get("durata_sec") or 0
        if dur_check == 0:
            return
        # Filtra: include la sessione SOLO se appartiene al progetto cercato
        if project_id or progetto:
            pid_sess  = (sessione.get("progetto_id") or "").strip()
            nome_sess = prog.lower()
            nome_cerca = (progetto or "").strip().lower()

            if pid_sess and project_id:
                # Sessione con ID: match esatto per ID
                if pid_sess != project_id:
                    return  # altro progetto
            elif nome_cerca:
                # Sessione senza ID (storico): confronta nome
                # Prova match esatto, poi parziale (nome tecnico dentro nome UI)
                if nome_sess == nome_cerca:
                    pass  # match esatto OK
                elif nome_cerca in nome_sess or nome_sess in nome_cerca:
                    pass  # un nome contiene l'altro OK
                else:
                    return  # nessun match
            else:
                # project_id presente ma pid_sess vuoto e progetto vuoto: escludi
                return
        dur = sessione.get("durata_sec") or 0
        if prog not in agg:
            agg[prog] = {
                "ore_sec":      0,
                "n_sessioni":   0,
                "prima_data":   sessione.get("data") or "",
                "ultima_data":  sessione.get("data") or "",
            }
        a = agg[prog]
        a["ore_sec"]    += max(0, dur)
        a["n_sessioni"] += 1
        if sessione.get("data"):
            if not a["prima_data"] or sessione["data"] < a["prima_data"]:
                a["prima_data"] = sessione["data"]
            if sessione["data"] > a["ultima_data"]:
                a["ultima_data"] = sessione["data"]

    for s in data.get("sessioni", []):
        _aggiungi(s)

    # Carica anche archivi annuali se presenti (es. lavorazioni_2025.json)
    # ESCLUDI lavorazioni_log.json che è già stato caricato sopra
    base = (config.get("tools_toa_folder") or
            config.get("percorso_nc_base") or ".")
    log_principale = Path(base) / "lavorazioni_log.json"
    for arch_path in sorted(Path(base).glob("lavorazioni_*.json")):
        if arch_path.resolve() == log_principale.resolve():
            continue  # già caricato — evita doppio conteggio
        try:
            arch = json.loads(arch_path.read_text(encoding="utf-8"))
            for s in arch.get("sessioni", []):
                _aggiungi(s)
        except Exception:
            pass

    # Aggiunge stringa ore formattata
    for prog, a in agg.items():
        a["ore_str"] = _durata_str(a["ore_sec"])

    if progetto or project_id:
        # Cerca per nome esatto
        entry = agg.get((progetto or "").strip()) if progetto else None
        if not entry and progetto:
            # Fallback: cerca case-insensitive o match parziale
            nome_norm = progetto.strip().lower()
            for k, v in agg.items():
                if k.lower() == nome_norm:
                    entry = v
                    break
        if not entry:
            return {"progetto": progetto or "", "ore_sec": 0, "ore_str": "00:00:00",
                    "n_sessioni": 0, "prima_data": None, "ultima_data": None}
        return {"progetto": progetto or "", **entry}

    # Ordina per ore decrescenti
    return {
        "progetti": dict(sorted(agg.items(), key=lambda x: -x[1]["ore_sec"])),
        "n_progetti": len(agg),
    }




@router.post("/reset-stato-corrente")
async def reset_stato_corrente():
    """
    Reset di emergenza dello stato_corrente bloccato.
    Chiude la sessione orfana e azzera sc mantenendo fermo_sec_giornaliero.
    """
    config = carica_configurazione()
    data = _load_log(config)
    sc = data.get("stato_corrente", {})
    now = _now_iso()

    # Chiudi sessione orfana se esiste
    if sc.get("sessione_id"):
        sess = _find_sess(data, sc["sessione_id"])
        if sess and sess.get("fine") is None:
            _chiudi_sessione(data, sc, sc.get("ultimo_tick") or now)

    # Preserva fermo giornaliero
    _fs = sc.get("fermo_sec_giornaliero", 0)
    _fd = sc.get("fermo_data", "")
    today = now[:10]

    sc.clear()
    if _fs and _fd == today:
        sc["fermo_sec_giornaliero"] = _fs
        sc["fermo_data"] = _fd

    _save_log(config, data)
    return {"ok": True, "msg": "stato_corrente resettato", "fermo_preservato": _fs}

@router.get("/debug-stato-corrente")
async def debug_stato_corrente():
    """Debug: espone stato_corrente raw del log."""
    config = carica_configurazione()
    data = _load_log(config)
    sc = data.get("stato_corrente", {})
    return {
        "stato_corrente": sc,
        "ultimo_tick_fermo": sc.get("ultimo_tick_fermo"),
        "in_esecuzione": sc.get("in_esecuzione"),
        "in_pausa": sc.get("in_pausa"),
        "programma_corrente": sc.get("programma_corrente"),
        "sessione_id": sc.get("sessione_id"),
        "fermo_sec_giornaliero": sc.get("fermo_sec_giornaliero"),
    }
@router.get("/debug-sessioni")
async def debug_sessioni_progetto(progetto: str, giorni: int = 365):
    """
    Debug: lista dettagliata di tutte le sessioni per un progetto.
    ?progetto=4297_0006&giorni=365
    Ritorna ogni sessione con inizio, fine, durata, programmi.
    """
    config  = carica_configurazione()
    data    = _load_log(config)
    cutoff  = (datetime.now() - timedelta(days=giorni)).strftime("%Y-%m-%d")
    nome    = progetto.strip()

    sessioni_trovate = []
    for s in data.get("sessioni", []):
        if (s.get("progetto") or "").strip() != nome:
            continue
        data_s = s.get("data") or ""
        if data_s < cutoff:
            continue
        dur = s.get("durata_sec")
        if dur is None:
            # Sessione aperta: usa somma programmi (non wall-clock che include fermi)
            pgms_chiusi = [p for p in s.get("programmi", []) if p.get("fine")]
            dur = sum(p.get("durata_sec") or 0 for p in pgms_chiusi)
        pgms = s.get("programmi", [])
        sessioni_trovate.append({
            "data":       data_s,
            "inizio":     s.get("inizio", "")[:19],
            "fine":       (s.get("fine") or "aperta")[:19],
            "durata_str": _durata_str(dur or 0),
            "durata_sec": dur or 0,
            "pallet":     s.get("pallet"),
            "n_programmi": len(pgms),
            "programmi":  [
                {
                    "filename":    p.get("filename", ""),
                    "inizio":      (p.get("inizio") or "")[:19],
                    "fine":        (p.get("fine") or "")[:19],
                    "durata_str":  _durata_str(p.get("durata_sec") or 0),
                    "durata_sec":  p.get("durata_sec") or 0,
                }
                for p in pgms
            ],
        })

    sessioni_trovate.sort(key=lambda s: s["data"])
    totale_sec = sum(s["durata_sec"] for s in sessioni_trovate)

    return {
        "progetto":      nome,
        "n_sessioni":    len(sessioni_trovate),
        "totale_str":    _durata_str(totale_sec),
        "totale_sec":    totale_sec,
        "sessioni":      sessioni_trovate,
    }


@router.get("/debug-ore-progetto")
async def debug_ore_progetto(progetto: str, project_id: str = None):
    """
    Debug: mostra esattamente quali sessioni vengono incluse/escluse
    nel calcolo ore-progetto, con motivazione.
    """
    config = carica_configurazione()
    data   = _load_log(config)
    sessioni = data.get("sessioni", [])

    righe = []
    tot = 0

    for s in sessioni:
        prog  = (s.get("progetto") or "").strip()
        fine  = s.get("fine") or ""
        dur   = s.get("durata_sec") or 0
        pid   = (s.get("progetto_id") or "").strip()
        inizio = (s.get("inizio") or "")[:16]
        fine16 = fine[:16] if fine and fine != "aperta" else "APERTA"

        # Determina motivo inclusione/esclusione
        if not prog or prog == "—":
            motivo = "SKIP: prog vuoto"
        elif not fine or fine == "aperta":
            motivo = "SKIP: sessione aperta"
        elif dur == 0:
            motivo = "SKIP: durata zero"
        elif project_id and pid and pid != project_id:
            motivo = f"SKIP: project_id mismatch ({pid})"
        elif progetto:
            nome_n = prog.lower()
            cerca  = progetto.strip().lower()
            if nome_n == cerca or cerca in nome_n or nome_n in cerca:
                tot += dur
                motivo = f"ADD: {dur//3600}h{(dur%3600)//60:02d}m"
            else:
                motivo = f"SKIP: nome '{prog}' != '{progetto}'"
        else:
            motivo = "SKIP: nessun filtro"

        righe.append({
            "inizio": inizio,
            "fine":   fine16,
            "prog":   prog,
            "dur":    f"{dur//3600}h{(dur%3600)//60:02d}m",
            "motivo": motivo,
        })

    return {
        "progetto":    progetto,
        "totale_str":  _durata_str(tot),
        "totale_sec":  tot,
        "n_sessioni_incluse": sum(1 for r in righe if r["motivo"].startswith("ADD")),
        "sessioni":    righe,
    }


@router.get("/debug-ore-progetto")
async def debug_ore_progetto(progetto: str, project_id: str = None):
    """Debug: mostra ogni sessione inclusa/esclusa nel calcolo ore-progetto."""
    config = carica_configurazione()
    data   = _load_log(config)
    righe  = []
    tot    = 0

    for s in data.get("sessioni", []):
        prog  = (s.get("progetto") or "").strip()
        fine  = s.get("fine") or ""
        dur   = s.get("durata_sec") or 0
        pid   = (s.get("progetto_id") or "").strip()
        inizio = (s.get("inizio") or "")[:16]
        fine16 = fine[:16] if fine and fine != "aperta" else "APERTA"

        if not prog or prog == "—":
            motivo = "SKIP-vuoto"
        elif not fine or fine == "aperta":
            motivo = "SKIP-aperta"
        elif dur == 0:
            motivo = "SKIP-zero"
        elif project_id and pid and pid != project_id:
            motivo = f"SKIP-id({pid})"
        else:
            nome_n = prog.lower()
            cerca  = progetto.strip().lower()
            if nome_n == cerca or cerca in nome_n or nome_n in cerca:
                tot += dur
                motivo = f"ADD {dur//3600}h{(dur%3600)//60:02d}m"
            else:
                motivo = f"SKIP-nome({prog})"

        righe.append({"inizio":inizio,"fine":fine16,"prog":prog,"motivo":motivo})

    return {
        "progetto":    progetto,
        "totale_str":  _durata_str(tot),
        "totale_sec":  tot,
        "n_incluse":   sum(1 for r in righe if r["motivo"].startswith("ADD")),
        "sessioni":    righe,
    }


@router.get("/rendiconto-progetto")
async def get_rendiconto_progetto(project_id: str):
    """
    Rendiconto completo di progetto: timeline, ore macchina, utensili, programmi.
    """
    from api.routers.progetti import _load_progetti, _load_deliveries
    config   = carica_configurazione()
    log_data = _load_log(config)

    # ── Carica progetto ──────────────────────────────────────────────────────
    proj_data = _load_progetti(config)
    progetto  = next((p for p in proj_data.get("projects", [])
                      if p.get("id") == project_id), None)
    if not progetto:
        raise HTTPException(status_code=404, detail="Progetto non trovato")

    nome    = progetto.get("name", "")
    creato  = progetto.get("createdAt", "")
    colore  = progetto.get("color", "#1D5FAD")

    # ── Delivery (scadenza / consegna) ───────────────────────────────────────
    deliveries = _load_deliveries(config)
    delivery   = next((d for d in deliveries if d.get("projectId") == project_id), None)
    scadenza       = delivery.get("dueDate")      if delivery else None
    consegnato     = delivery.get("delivered")    if delivery else False
    consegnato_at  = delivery.get("deliveredAt")  if delivery else None

    # ── Sessioni di lavorazione ──────────────────────────────────────────────
    # Filtri sui nomi file di sistema — esclude SPF, CMA, ecc.
    # NON include /_N_WKS che è il prefisso del path grezzo, non del filename estratto
    FILTRI = ("_N_CMA", "_N_CST", "_N_SYF", "_N_MPF_DIR", "BPOSAXIS",
              "_SPF", "0_MAIN_", "PALLET5", "ROTTURA", "LASEREIN", "BLMEAS",
              "FINE_PALLET", "BAX2START")

    sessioni_proj = []
    for s in log_data.get("sessioni", []):
        pid = (s.get("progetto_id") or "").strip()
        pnm = (s.get("progetto")    or "").strip()
        # Filtra per project_id se disponibile (più affidabile)
        if pid:
            if pid != project_id:
                continue
        else:
            # Fallback: confronto nome progetto
            if pnm.lower() != nome.lower():
                continue
        # Verifica di coerenza: almeno un programma deve contenere
        # la commessa nel nome (evita sessioni orfane con project_id sbagliato)
        pgms = s.get("programmi", [])
        if pgms:
            commessa_prefix = nome.split("_")[0]  # es. "4350"
            has_matching_pgm = any(
                commessa_prefix in (p.get("filename") or "").upper()
                for p in pgms
            )
            if not has_matching_pgm:
                continue
        sessioni_proj.append(s)

    # ── Aggregazione programmi ───────────────────────────────────────────────
    # Prefissi validi per questa commessa — es. "4297" dalla commessa "4297_0006"
    commessa_prefissi = set()
    commessa_prefissi.add(nome.split("_")[0].upper())  # prefisso numerico principale
    commessa_prefissi.add(nome.upper())                  # nome completo commessa

    pgm_agg: dict = {}
    for s in sessioni_proj:
        for p in s.get("programmi", []):
            fn  = (p.get("filename") or "").strip()
            if not fn or any(f in fn.upper() for f in FILTRI):
                continue
            # Escludi programmi wrapper Sinumerik (path _N_ senza .MPF finale)
            # es: /_N_WKS_DIR/_N_4350_0221_WPD/_N_4350_0221_01_020_MP (senza F)
            fn_upper = fn.upper()
            if fn_upper.startswith("/_N_") and not fn_upper.endswith("_MPF"):
                continue
            # Escludi programmi di altre commesse — il filename deve contenere
            # almeno uno dei prefissi della commessa corrente
            fn_base = fn_upper.replace(".MPF", "").split("/")[-1]
            if fn_base.startswith("_N_"):
                fn_base = fn_base[3:]
            if not any(fn_base.startswith(pfx) for pfx in commessa_prefissi):
                continue
            dur = p.get("durata_sec") or 0
            fn_key = fn.upper().replace(".MPF", "").split("/")[-1]
            # Rimuovi prefisso _N_ se presente
            if fn_key.startswith("_N_"):
                fn_key = fn_key[3:]
            if fn_key not in pgm_agg:
                pgm_agg[fn_key] = {"filename": fn_key, "durata_sec": 0, "n_esecuzioni": 0}
            pgm_agg[fn_key]["durata_sec"]   += dur
            pgm_agg[fn_key]["n_esecuzioni"] += 1

    # tempoStimato CAM per confronto
    for s in progetto.get("steps", []):
        for t in s.get("tasks", []):
            if t.get("text", "").strip().lower() != "fresatura":
                continue
            for pg in t.get("programs", []):
                fn_key = (pg.get("filename") or "").upper().replace(".MPF", "").split("/")[-1]
                if fn_key in pgm_agg:
                    stima = pg.get("tempoStimato")
                    if stima:
                        try:
                            pgm_agg[fn_key]["stima_sec"] = int(stima) * 60
                        except Exception:
                            pass

    programmi_list = sorted(pgm_agg.values(), key=lambda x: -x["durata_sec"])

    # ── Aggregazione utensili ────────────────────────────────────────────────
    # Struttura log: {"alias": durata_sec_int} — dict piatto
    ut_agg: dict = {}
    for s in sessioni_proj:
        u = s.get("utensili") or {}
        if not isinstance(u, dict):
            continue
        for alias, val in u.items():
            if alias not in ut_agg:
                ut_agg[alias] = {"alias": alias, "durata_sec": 0}
            if isinstance(val, (int, float)):
                ut_agg[alias]["durata_sec"] += int(val)
            elif isinstance(val, dict):
                ut_agg[alias]["durata_sec"] += val.get("durata_sec", 0) or 0

    utensili_list = sorted(ut_agg.values(), key=lambda x: -x["durata_sec"])
    for u in utensili_list:
        u["ore_str"] = _durata_str(u["durata_sec"])

    # ── KPI principali ───────────────────────────────────────────────────────
    ore_macchina_sec = sum(
        s.get("durata_sec") or sum(p.get("durata_sec") or 0 for p in s.get("programmi", []))
        for s in sessioni_proj
    )
    n_sessioni  = len(sessioni_proj)
    n_programmi = len(pgm_agg)
    n_utensili  = len(ut_agg)

    # Fasi di lavorazione (steps con almeno una sessione)
    fasi = [s.get("title", "") for s in progetto.get("steps", [])]

    # Conteggio programmi dal progetto
    pgm_progetto = [
        pg for s in progetto.get("steps", [])
        for t in s.get("tasks", [])
        if t.get("text", "").strip().lower() == "fresatura"
        for pg in t.get("programs", [])
        if pg.get("tipoGruppo") != "ipm"
    ]
    n_pgm_totali    = len(pgm_progetto)
    n_pgm_completati = sum(1 for pg in pgm_progetto if pg.get("stato") == "completato")
    stima_tot_sec = sum(
        int(pg.get("tempoStimato") or 0) * 60
        for pg in pgm_progetto
        if pg.get("tempoStimato") and str(pg.get("tempoStimato")).strip() not in ("", "0")
    )

    # ── Timeline ─────────────────────────────────────────────────────────────
    prima_sess = sessioni_proj[0].get("inizio", "")[:10]  if sessioni_proj else None
    ultima_fine = None
    for s in reversed(sessioni_proj):
        f = s.get("fine")
        if f and f != "aperta":
            ultima_fine = f[:10]
            break

    # Giorni calendario progetto
    try:
        da = datetime.fromisoformat(creato)
        a  = datetime.fromisoformat(consegnato_at.replace("/","").replace(" ","T")[:10])              if consegnato_at else datetime.now()
        giorni_totali = (a - da).days
    except Exception:
        giorni_totali = None

    try:
        giorni_macchina = (
            datetime.fromisoformat(ultima_fine) -
            datetime.fromisoformat(prima_sess)
        ).days + 1 if prima_sess and ultima_fine else None
    except Exception:
        giorni_macchina = None

    # ── Sessioni per il dettaglio timeline ───────────────────────────────────
    sessioni_out = []
    for s in sessioni_proj:
        dur = s.get("durata_sec") or sum(p.get("durata_sec") or 0 for p in s.get("programmi",[]))
        if dur == 0:
            continue
        pgms_validi = [
            p for p in s.get("programmi", [])
            if not any(f in (p.get("filename") or "").upper() for f in FILTRI)
            and (p.get("durata_sec") or 0) >= 30
        ]
        sessioni_out.append({
            "data":        s.get("data", ""),
            "inizio":      s.get("inizio", "")[:19],
            "fine":        (s.get("fine") or "")[:19],
            "durata_sec":  dur,
            "durata_str":  _durata_str(dur),
            "n_programmi": len(pgms_validi),
        })

    # ── Ore CAM (da cam_tracker_data.json) ───────────────────────────────────
    ore_cam_sec = 0
    ore_cam_per_op: list[dict] = []
    try:
        cam_file = Path("cam_tracker_data.json")
        if cam_file.exists():
            import json as _json
            cam_data = _json.loads(cam_file.read_text(encoding="utf-8"))
            # nome = es. "4297_0006"
            # Cerca prima per project esatto (es. project="4297_0006")
            # Solo se non trova nulla, cerca per commessa+operazione
            commessa_proj = nome.split("_")[0] if "_" in nome else nome
            operazione_proj = nome.split("_")[1] if "_" in nome else ""

            # Match esatto project_id
            cam_entries = [e for e in cam_data
                           if e.get("project", "").upper() == nome.upper()]

            # Fallback: commessa + operazione (es. commessa=4297, op=0006)
            if not cam_entries and operazione_proj:
                cam_entries = [
                    e for e in cam_data
                    if e.get("commessa", "").upper() == commessa_proj.upper()
                    and (e.get("operazione") or "").lstrip("0") == operazione_proj.lstrip("0")
                ]

            ore_cam_sec = sum(e.get("seconds", 0) for e in cam_entries)
            op_agg: dict[str, int] = {}
            for e in cam_entries:
                op = e.get("operazione") or e.get("project", "")
                op_agg[op] = op_agg.get(op, 0) + e.get("seconds", 0)
            ore_cam_per_op = [
                {"operazione": op, "seconds": sec, "hours": round(sec / 3600, 2)}
                for op, sec in sorted(op_agg.items(), key=lambda x: -x[1])
            ]
    except Exception as _e:
        pass  # cam_tracker non ancora configurato — non blocca il rendiconto

    return {
        "progetto": {
            "id":    project_id,
            "nome":       nome,
            "descrizione": progetto.get("description") or "",
            "colore":     colore,
            "creato":     creato,
        },
        "delivery": {
            "scadenza":      scadenza,
            "consegnato":    consegnato,
            "consegnato_at": consegnato_at,
        },
        "timeline": {
            "apertura_progetto": creato,
            "inizio_macchina":   prima_sess,
            "fine_macchina":     ultima_fine,
            "consegna":          consegnato_at,
            "scadenza":          scadenza,
            "giorni_totali":     giorni_totali,
            "giorni_macchina":   giorni_macchina,
        },
        "kpi": {
            "ore_macchina_sec":  ore_macchina_sec,
            "ore_macchina_str":  _durata_str(ore_macchina_sec),
            "ore_cam_sec":       ore_cam_sec,
            "ore_cam_str":       _durata_str(ore_cam_sec),
            "ore_cam_per_op":    ore_cam_per_op,
            "n_sessioni":        n_sessioni,
            "n_programmi_eseguiti": n_programmi,
            "n_programmi_totali":   n_pgm_totali,
            "n_programmi_completati": n_pgm_completati,
            "n_utensili":        n_utensili,
            "n_fasi":            len(fasi),
            "fasi":              fasi,
            "stima_tot_sec":     stima_tot_sec,
            "stima_tot_str":     _durata_str(stima_tot_sec),
            "scostamento_pct":   round((ore_macchina_sec - stima_tot_sec) / stima_tot_sec * 100, 1)
                                 if stima_tot_sec > 0 else None,
        },
        "programmi":  programmi_list,
        "utensili":   utensili_list,
        "sessioni":   sessioni_out,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYTICS AVANZATE
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/analytics-commesse")
async def get_analytics_commesse():
    """
    Analytics avanzate per tutte le commesse:
    - Ratio CAM/Macchina per commessa
    - OEE giornaliero (ultimi 30gg)
    - Stima fine lavori per commesse aperte
    - Alert scadenze a rischio
    """
    from api.routers.progetti import _load_progetti, _load_deliveries
    from datetime import datetime, date, timedelta

    config   = carica_configurazione()
    log_data = _load_log(config)
    proj_data = _load_progetti(config)
    deliveries = _load_deliveries(config)
    progetti = proj_data.get("projects", [])

    # ── Pallet state per Gantt ──────────────────────────────────────────────
    pallet_map: dict[str, dict] = {}  # progetto_id → {numero, stato}
    try:
        from api.routers.pallet import _load as _load_pallet, _pallet_path
        ps = _load_pallet(config)
        for pal in ps.get("pallet", []):
            pid = pal.get("progetto_id")
            if pid:
                pallet_map[pid] = {
                    "numero": pal.get("numero"),
                    "stato":  pal.get("stato", ""),
                }
    except Exception:
        pass

    # ── Ore macchina per commessa (da log sessioni) ──────────────────────────
    ore_mac: dict[str, float] = {}
    for s in log_data.get("sessioni", []):
        nome = (s.get("progetto") or "").strip()
        if not nome or nome == "—":
            continue
        dur = s.get("durata_sec") or sum(
            p.get("durata_sec", 0) for p in s.get("programmi", [])
        )
        ore_mac[nome] = ore_mac.get(nome, 0) + dur

    # ── Ore CAM per commessa (da cam_tracker_data.json) ─────────────────────
    ore_cam: dict[str, float] = {}
    cam_per_commessa: dict[str, list] = {}
    try:
        cam_file = Path("cam_tracker_data.json")
        if cam_file.exists():
            cam_data = json.loads(cam_file.read_text(encoding="utf-8"))
            for e in cam_data:
                proj_id = (e.get("project") or "").upper()
                commessa = (e.get("commessa") or proj_id.split("_")[0]).upper()
                ore_cam[proj_id] = ore_cam.get(proj_id, 0) + e.get("seconds", 0)
                cam_per_commessa.setdefault(commessa, []).append(e)
    except Exception:
        pass

    # ── OEE ultimi 30 giorni ─────────────────────────────────────────────────
    # Disponibilità = ore macchina attiva / ore turno (ipotesi 8h/giorno)
    ORE_TURNO_SEC = 24 * 3600
    oggi = date.today()
    oee_giorni = []
    sessioni_all = log_data.get("sessioni", [])
    for i in range(29, -1, -1):
        giorno = (oggi - timedelta(days=i)).isoformat()
        sess_g = [s for s in sessioni_all if s.get("data") == giorno]
        ore_att = sum(
            (s.get("durata_sec") or sum(p.get("durata_sec", 0)
             for p in s.get("programmi", [])))
            for s in sess_g
        )
        # Prestazione: cicli reali vs teorici (usa media cicli se disponibile)
        n_pgm = sum(len(s.get("programmi", [])) for s in sess_g)
        idx = log_data.get("cicli_utensile", {})
        cicli_teorici = 0
        cicli_reali = 0
        for s in sess_g:
            for p in s.get("programmi", []):
                fname = (p.get("filename") or "").upper()
                entry = idx.get(fname)
                if entry and entry.get("media"):
                    cicli_teorici += entry["media"]
                    cicli_reali   += p.get("durata_sec") or entry["media"]
        prestazione = round(cicli_teorici / cicli_reali, 3) if cicli_reali > 0 else None
        disponibilita = round(min(ore_att / ORE_TURNO_SEC, 1.0), 3) if ore_att > 0 else 0
        oee_val = round(disponibilita * (prestazione or 1.0), 3) if ore_att > 0 else 0
        if ore_att > 0:
            oee_giorni.append({
                "data":           giorno,
                "disponibilita":  round(disponibilita * 100, 1),
                "prestazione":    round((prestazione or 1.0) * 100, 1),
                "oee_pct":        round(oee_val * 100, 1),
                "ore_attive_sec": ore_att,
                "n_programmi":    n_pgm,
            })

    oee_medio = round(
        sum(g["oee_pct"] for g in oee_giorni) / len(oee_giorni), 1
    ) if oee_giorni else None

    # ── Fattore di calibrazione K globale ────────────────────────────────────
    analytics_progetti = []
    delivery_map = {d.get("projectId"): d for d in deliveries}

    k_campioni = []
    K_MIN = 0.5   # minimo plausibile: 1h CAM → almeno 30min macchina
    K_MAX = 20.0  # massimo plausibile: 1h CAM → max 20h macchina
    CAM_MIN_SEC = 1800  # almeno 30 min CAM per essere un campione valido
    for p in progetti:
        nome_p = p.get("name", "")
        delivery_p = delivery_map.get(p.get("id", ""), {})
        if not delivery_p.get("delivered"):
            continue
        mac_s = ore_mac.get(nome_p, 0)
        cam_s = ore_cam.get(nome_p.upper(), 0)
        if mac_s > 0 and cam_s >= CAM_MIN_SEC:
            k = mac_s / cam_s
            if K_MIN <= k <= K_MAX:
                k_campioni.append(k)

    k_medio   = round(sum(k_campioni) / len(k_campioni), 2) if k_campioni else None
    k_std     = round((sum((k - k_medio)**2 for k in k_campioni) / len(k_campioni))**0.5, 2) if len(k_campioni) > 1 else None
    n_campioni_k = len(k_campioni)

    # Confidenza: bassa <3, media 3-7, alta ≥8 commesse complete con dati CAM
    if n_campioni_k >= 8:
        k_confidenza = "alta"
    elif n_campioni_k >= 3:
        k_confidenza = "media"
    elif n_campioni_k >= 1:
        k_confidenza = "bassa"
    else:
        k_confidenza = None

    # ── Ratio CAM/Macchina e stima fine lavori ───────────────────────────────
    for p in progetti:
        nome = p.get("name", "")
        pid  = p.get("id", "")
        commessa = nome.split("_")[0] if "_" in nome else nome

        ore_mac_sec = ore_mac.get(nome, 0)
        ore_cam_sec = ore_cam.get(nome.upper(), 0)

        # Ratio CAM/Macchina (grezzo, senza correzione)
        ratio = round(ore_mac_sec / ore_cam_sec, 2) if ore_cam_sec > 0 else None

        # Stima fine lavori
        # Programmi rimanenti × tempo ciclo medio
        # Stima rimanente TOTALE (tutti i non-completati) — per giorni_rimanenti
        pgm_rimanenti = [
            pg for s in p.get("steps", [])
            for t in s.get("tasks", [])
            if t.get("text", "").strip().lower() == "fresatura"
            for pg in t.get("programs", [])
            if pg.get("tipoGruppo") != "ipm" and pg.get("stato") != "completato"
        ]
        stima_rimanente_sec = sum(
            int(pg.get("tempoStimato") or 0) * 60
            for pg in pgm_rimanenti
            if pg.get("tempoStimato")
        )

        # Stima in_main — solo programmi pianificati nel MAIN (per Gantt e ore_rimanenti)
        pgm_in_main = [
            pg for s in p.get("steps", [])
            for t in s.get("tasks", [])
            if t.get("text", "").strip().lower() == "fresatura"
            for pg in t.get("programs", [])
            if pg.get("tipoGruppo") != "ipm"
            and pg.get("stato") in ("in_main", "in_lavorazione", "in_macchina")
        ]
        stima_in_main_sec = sum(
            int(pg.get("tempoStimato") or 0) * 60
            for pg in pgm_in_main
            if pg.get("tempoStimato")
        )

        # Applica fattore K alla stima CAM se disponibile
        # Senza K: stima grezza (include bias sistematico)
        # Con K: stima corretta basata su storico reale
        if stima_rimanente_sec > 0 and k_medio and k_medio <= K_MAX:
            stima_rimanente_corretta_sec = round(stima_rimanente_sec * k_medio)
        else:
            stima_rimanente_corretta_sec = stima_rimanente_sec
        # Usa ore/giorno reali degli ultimi 7gg per questo progetto
        ultimi7 = [
            s for s in sessioni_all
            if (s.get("progetto") or "") == nome
            and s.get("data", "") >= (oggi - timedelta(days=7)).isoformat()
        ]
        ore_giorno_reali = (
            sum(s.get("durata_sec") or 0 for s in ultimi7) / min(7, max(1, len(ultimi7)))
        ) if ultimi7 else 0

        # Se la media giornaliera è < 1h (progetto appena iniziato o pochissimi dati),
        # usa il fallback 60% del turno per evitare stime assurde (es. 5000gg)
        ORE_MINIMA_GIORNO = 3600  # 1h minima per una stima sensata
        if ore_giorno_reali < ORE_MINIMA_GIORNO:
            ore_giorno_reali = ORE_TURNO_SEC * 0.6  # default 60% utilizzo

        giorni_rimanenti = (
            round(stima_rimanente_corretta_sec / ore_giorno_reali)
            if ore_giorno_reali > 0 and stima_rimanente_corretta_sec > 0
            else None
        )
        # Cap a 365gg — oltre non ha senso mostrare una stima
        if giorni_rimanenti is not None and giorni_rimanenti > 365:
            giorni_rimanenti = None
        data_fine_stimata = (
            (oggi + timedelta(days=giorni_rimanenti)).isoformat()
            if giorni_rimanenti is not None and giorni_rimanenti <= 365
            else None
        )

        # ── Intervallo di confidenza ±N giorni ──────────────────────────────
        # Basato sulla varianza delle ore/giorno negli ultimi 14gg (reparto)
        # Varianza alta → ±più giorni; varianza bassa → ±meno giorni
        confidenza_giorni = None
        confidenza_label  = None
        if giorni_rimanenti is not None and ore_giorno_reali > 0:
            # Raccoglie ore/giorno degli ultimi 14gg su TUTTI i progetti (velocità reparto)
            campioni_14gg = []
            for s in sessioni_all:
                if s.get("data", "") >= (oggi - timedelta(days=14)).isoformat():
                    dur = s.get("durata_sec") or 0
                    if dur > 0:
                        campioni_14gg.append(dur)

            if len(campioni_14gg) >= 3:
                import statistics as _stats
                media_camp = _stats.mean(campioni_14gg)
                stdev_camp = _stats.stdev(campioni_14gg) if len(campioni_14gg) >= 2 else 0
                cv = stdev_camp / media_camp if media_camp > 0 else 0  # coefficiente di variazione

                # cv basso (<0.3) → alta confidenza ±1-2gg
                # cv medio (0.3-0.6) → media confidenza ±3-5gg
                # cv alto (>0.6) → bassa confidenza ±7+gg
                if stima_rimanente_corretta_sec > 0:
                    # Propaga l'incertezza: delta_giorni = cv × giorni_rimanenti
                    confidenza_giorni = max(1, round(cv * giorni_rimanenti))
                    confidenza_giorni = min(confidenza_giorni, 30)  # cap a 30gg
                    if cv < 0.3:
                        confidenza_label = "alta"
                    elif cv < 0.6:
                        confidenza_label = "media"
                    else:
                        confidenza_label = "bassa"
            else:
                # Pochi dati storici → confidenza bassa con ±fisso
                confidenza_giorni = min(max(2, round(giorni_rimanenti * 0.3)), 14)
                confidenza_label  = "bassa"

        # Alert scadenza
        delivery = delivery_map.get(pid, {})
        scadenza = delivery.get("dueDate")
        consegnato = delivery.get("delivered", False)
        alert = None
        if not consegnato and scadenza and data_fine_stimata:
            giorni_scarto = (
                date.fromisoformat(scadenza) -
                date.fromisoformat(data_fine_stimata)
            ).days
            if giorni_scarto < 0:
                alert = {"tipo": "ritardo", "giorni": abs(giorni_scarto),
                         "msg": f"Stima fine lavori {abs(giorni_scarto)}gg oltre scadenza"}
            elif giorni_scarto <= 3:
                alert = {"tipo": "warning", "giorni": giorni_scarto,
                         "msg": f"Solo {giorni_scarto}gg di margine sulla scadenza"}

        if ore_mac_sec > 0 or ore_cam_sec > 0:
            analytics_progetti.append({
                "id":                  pid,
                "nome":                nome,
                "commessa":            commessa,
                "colore":              p.get("color", "#1D5FAD"),
                "ore_macchina_sec":    ore_mac_sec,
                "ore_macchina_str":    _durata_str(ore_mac_sec),
                "ore_cam_sec":         ore_cam_sec,
                "ore_cam_str":         _durata_str(ore_cam_sec),
                "ratio_cam_macchina":  ratio,
                "stima_rimanente_sec": stima_rimanente_sec,
                "stima_rimanente_str": _durata_str(stima_rimanente_sec),
                "stima_corretta_sec":  stima_rimanente_corretta_sec,
                "stima_corretta_str":  _durata_str(stima_rimanente_corretta_sec),
                "k_applicato":         k_medio,
                "giorni_rimanenti":    giorni_rimanenti,
                "data_fine_stimata":   data_fine_stimata,
                "confidenza_giorni":   confidenza_giorni,
                "confidenza_label":    confidenza_label,
                "scadenza":            scadenza,
                "consegnato":          consegnato,
                "alert":               alert,
                "pallet_numero":       pallet_map.get(pid, {}).get("numero"),
                "is_live":             pallet_map.get(pid, {}).get("stato") == "in_lavorazione",
                "ore_rimanenti":       round((stima_rimanente_corretta_sec or stima_rimanente_sec or 0) / 3600, 2),
            })

    # Ordina: alert prima, poi per ore macchina decrescenti
    analytics_progetti.sort(key=lambda x: (
        0 if x.get("alert", {}) and x["alert"]["tipo"] == "ritardo" else
        1 if x.get("alert", {}) and x["alert"]["tipo"] == "warning" else 2,
        -x["ore_macchina_sec"]
    ))

    return {
        "oee": {
            "medio_pct":  oee_medio,
            "giorni":     oee_giorni[-14:],
            "n_giorni":   len(oee_giorni),
        },
        "calibrazione": {
            "k_medio":    k_medio,
            "k_std":      k_std,
            "n_campioni": n_campioni_k,
            "confidenza": k_confidenza,
            "nota": (
                f"K = {k_medio:.2f} — per ogni ora CAM stimi {k_medio:.1f}h macchina in media "
                f"({n_campioni_k} commesse complete, confidenza {k_confidenza})"
                if k_medio else
                "Nessun campione — serve almeno 1 commessa completata con dati CAM e macchina"
            ),
        },
        "progetti":   analytics_progetti,
        "n_con_cam":  sum(1 for p in analytics_progetti if p["ore_cam_sec"] > 0),
        "n_alert":    sum(1 for p in analytics_progetti if p.get("alert")),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ALERT UTENSILI PREDITTIVI
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/alert-utensili")
async def get_alert_utensili():
    """
    Analisi predittiva utensili — rileva degrado in corso.
    Criteri di alert:
    - slope > 2.0 sec/ciclo per almeno 3 cicli consecutivi
    - CV% > 20% (processo instabile)
    - durata ultimo ciclo > media + 2.5σ
    """
    config = carica_configurazione()
    data   = _load_log(config)
    idx    = data.get("cicli_utensile", {})

    alerts = []
    warnings = []

    for chiave, entry in idx.items():
        if "|" not in chiave:
            continue  # skip per-filename, solo coppie utensile|filename
        utensile, filename = chiave.split("|", 1)
        campioni = entry.get("campioni", [])
        n = len(campioni)
        if n < 3:
            continue

        media  = entry.get("media", 0)
        std    = entry.get("std", 0)
        slope  = entry.get("slope", 0) if hasattr(entry, "get") else 0

        # Ricalcola slope se non presente
        if "slope" not in entry and n >= 3:
            xs = list(range(n))
            x_mean = (n-1)/2
            slope_num = sum((xs[i]-x_mean)*(campioni[i]-media) for i in range(n))
            slope_den = sum((x-x_mean)**2 for x in xs)
            slope = round(slope_num/slope_den, 2) if slope_den else 0

        ultimo = campioni[-1] if campioni else 0
        cv_pct = round(std/media*100, 1) if media > 0 else 0
        delta_ultimo = round(ultimo - media) if media > 0 else 0
        soglia_alert = media + 2.5*std

        item = {
            "utensile":    utensile,
            "programma":   filename,
            "n_cicli":     n,
            "media_sec":   media,
            "std_sec":     std,
            "cv_pct":      cv_pct,
            "slope":       slope,
            "ultimo_sec":  ultimo,
            "delta_sec":   delta_ultimo,
            "soglia_sec":  round(soglia_alert),
        }

        if slope > 2.0 and n >= 5:
            item["tipo"] = "degrado"
            item["msg"]  = f"Ciclo in aumento +{slope:.1f}s/ciclo su {n} campioni"
            item["severita"] = "alta" if slope > 5.0 else "media"
            alerts.append(item)
        elif ultimo > soglia_alert and n >= 5:
            item["tipo"] = "picco"
            item["msg"]  = f"Ultimo ciclo {delta_ultimo}s oltre la media (soglia +{round(2.5*std)}s)"
            item["severita"] = "alta"
            alerts.append(item)
        elif cv_pct > 20:
            item["tipo"] = "instabile"
            item["msg"]  = f"Processo instabile CV={cv_pct}% (soglia 20%)"
            item["severita"] = "bassa"
            warnings.append(item)

    # Ordina per severità e slope
    alerts.sort(key=lambda x: (
        0 if x["severita"] == "alta" else 1,
        -abs(x.get("slope", 0))
    ))

    return {
        "alert":    alerts,
        "warning":  warnings,
        "n_alert":  len(alerts),
        "n_warning": len(warnings),
        "totale_utensili_monitorati": sum(1 for k in idx if "|" in k),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURAZIONE NOTIFICHE
# ═══════════════════════════════════════════════════════════════════════════════

_NOTIF_FILE = Path("notifiche_config.json")

def _load_notif_config() -> dict:
    if _NOTIF_FILE.exists():
        try:
            return json.loads(_NOTIF_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {
        "email": {"attivo": False, "smtp_host": "", "smtp_port": 587,
                  "mittente": "", "password": "", "destinatari": []},
        "webhook": {"attivo": False, "url": "", "secret": ""},
        "soglie": {"slope_alert": 2.0, "cv_warning": 20.0,
                   "giorni_scadenza_warning": 3, "idle_min": 5},
    }

@router.get("/notifiche/config")
async def get_notif_config():
    cfg = _load_notif_config()
    # Oscura password
    if cfg.get("email", {}).get("password"):
        cfg["email"]["password"] = "••••••••"
    return cfg

@router.post("/notifiche/config")
async def save_notif_config(body: dict = Body(...)):
    cfg = _load_notif_config()
    # Merge preservando password se oscurata
    for sezione, valori in body.items():
        if sezione not in cfg:
            cfg[sezione] = {}
        for k, v in valori.items():
            if k == "password" and v == "••••••••":
                continue  # non sovrascrivere con il placeholder
            cfg[sezione][k] = v
    _NOTIF_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"ok": True}

@router.post("/notifiche/test")
async def test_notifica(body: dict = Body(...)):
    """Invia notifica di test via email o webhook."""
    cfg = _load_notif_config()
    tipo = body.get("tipo", "webhook")

    if tipo == "webhook":
        wcfg = cfg.get("webhook", {})
        if not wcfg.get("url"):
            return {"ok": False, "errore": "URL webhook non configurato"}
        try:
            import urllib.request, json as _json
            payload = _json.dumps({
                "tipo": "test",
                "msg": "DMGDesk — notifica di test",
                "ts": datetime.now().isoformat(),
            }).encode()
            req = urllib.request.Request(
                wcfg["url"], data=payload,
                headers={"Content-Type": "application/json"}, method="POST"
            )
            with urllib.request.urlopen(req, timeout=5) as resp:
                return {"ok": True, "status": resp.status}
        except Exception as e:
            return {"ok": False, "errore": str(e)}

    elif tipo == "email":
        ecfg = cfg.get("email", {})
        if not ecfg.get("smtp_host") or not ecfg.get("destinatari"):
            return {"ok": False, "errore": "Email non configurata"}
        try:
            import smtplib
            from email.mime.text import MIMEText
            msg = MIMEText("DMGDesk — notifica di test")
            msg["Subject"] = "DMGDesk Test Notifica"
            msg["From"] = ecfg["mittente"]
            msg["To"]   = ", ".join(ecfg["destinatari"])
            with smtplib.SMTP(ecfg["smtp_host"], ecfg.get("smtp_port", 587)) as s:
                s.starttls()
                s.login(ecfg["mittente"], ecfg.get("password", ""))
                s.sendmail(ecfg["mittente"], ecfg["destinatari"], msg.as_string())
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "errore": str(e)}

    return {"ok": False, "errore": "Tipo non supportato"}


async def _invia_notifica_alert(alert_list: list, tipo: str):
    """Chiamato automaticamente quando si rilevano nuovi alert."""
    if not alert_list:
        return
    cfg = _load_notif_config()

    msg_lines = [f"DMGDesk Alert — {tipo}", ""]
    for a in alert_list[:5]:
        msg_lines.append(f"• {a.get('utensile','?')} / {a.get('programma','?')}: {a.get('msg','')}")
    if len(alert_list) > 5:
        msg_lines.append(f"... e altri {len(alert_list)-5} alert")
    msg_text = "\n".join(msg_lines)

    # Webhook
    wcfg = cfg.get("webhook", {})
    if wcfg.get("attivo") and wcfg.get("url"):
        try:
            import urllib.request, json as _json
            payload = _json.dumps({
                "tipo": tipo, "n_alert": len(alert_list),
                "msg": msg_text, "ts": datetime.now().isoformat(),
                "alert": alert_list[:10],
            }).encode()
            req = urllib.request.Request(
                wcfg["url"], data=payload,
                headers={"Content-Type": "application/json"}, method="POST"
            )
            urllib.request.urlopen(req, timeout=5)
        except Exception:
            pass

    # Email
    ecfg = cfg.get("email", {})
    if ecfg.get("attivo") and ecfg.get("smtp_host") and ecfg.get("destinatari"):
        try:
            import smtplib
            from email.mime.text import MIMEText
            msg = MIMEText(msg_text)
            msg["Subject"] = f"DMGDesk Alert: {tipo} ({len(alert_list)})"
            msg["From"] = ecfg["mittente"]
            msg["To"]   = ", ".join(ecfg["destinatari"])
            with smtplib.SMTP(ecfg["smtp_host"], ecfg.get("smtp_port", 587)) as s:
                s.starttls()
                s.login(ecfg["mittente"], ecfg.get("password", ""))
                s.sendmail(ecfg["mittente"], ecfg["destinatari"], msg.as_string())
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT RENDICONTO JSON (base per PDF frontend)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/rendiconto-export/{project_id}")
async def export_rendiconto(project_id: str):
    """
    Rendiconto completo in formato ottimizzato per export PDF.
    Stessi dati di rendiconto-progetto con aggiunta di:
    - riepilogo_testo: paragrafi pronti per stampa
    - statistiche_utensili: aggregato per alias
    """
    from api.routers.progetti import _load_progetti, _load_deliveries
    config   = carica_configurazione()
    log_data = _load_log(config)
    proj_data = _load_progetti(config)
    progetto  = next((p for p in proj_data.get("projects", [])
                      if p.get("id") == project_id), None)
    if not progetto:
        raise HTTPException(status_code=404, detail="Progetto non trovato")

    # Riusa rendiconto-progetto richiamando direttamente la funzione
    # (evita duplicazione codice — lo richiama via HTTP interno)
    import httpx
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            f"http://localhost:8000/api/report/rendiconto-progetto?project_id={project_id}"
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code)
        rend = resp.json()

    kpi  = rend.get("kpi", {})
    tl   = rend.get("timeline", {})
    nome = rend["progetto"]["nome"]

    # Riepilogo testuale per PDF
    righe = [
        f"Commessa: {nome}",
        f"Ore lavorazione CNC: {kpi.get('ore_macchina_str','—')}",
        f"Ore programmazione CAM: {kpi.get('ore_cam_str','0h')}",
    ]
    if kpi.get("ore_cam_sec", 0) > 0 and kpi.get("ore_macchina_sec", 0) > 0:
        ratio = round(kpi["ore_macchina_sec"] / kpi["ore_cam_sec"], 1)
        righe.append(f"Ratio CAM/Macchina: 1 : {ratio}")
    if tl.get("giorni_totali"):
        righe.append(f"Durata commessa: {tl['giorni_totali']} giorni")
    if kpi.get("scostamento_pct") is not None:
        sc = kpi["scostamento_pct"]
        righe.append(f"Scostamento vs stima: {'+' if sc>0 else ''}{sc}%")

    rend["riepilogo_testo"] = righe
    rend["export_ts"] = datetime.now().isoformat()
    return rend


# ── Mappa termica utilizzo macchina ───────────────────────────────────────────
@router.get("/heatmap-utilizzo")
async def heatmap_utilizzo():
    """
    Ritorna una griglia giorni × fasce orarie con ore di produzione.
    Fasce: mattina 07:30-16:30, sera 16:30-22:00, notte 22:00-07:30
    Storico: ultime 4 settimane (28 giorni)
    """
    from datetime import date, timedelta, datetime as dt

    config = carica_configurazione()

    FASCE = [
        {"key": "notte",   "label": "Notte",   "da": 22*60,      "a": 24*60 + 7*60+30},  # 22:00-07:30 (+giorno)
        {"key": "mattina", "label": "Mattina",  "da": 7*60+30,    "a": 16*60+30},          # 07:30-16:30
        {"key": "sera",    "label": "Sera",     "da": 16*60+30,   "a": 22*60},             # 16:30-22:00
    ]
    # Valore assoluto massimo per cella = 8h (mattina) / 5.5h (sera) / 9.5h (notte)
    # Usiamo 8h come cap comune per normalizzare il colore
    MAX_SEC_CELLA = 8 * 3600

    oggi = date.today()
    da_data = oggi - timedelta(days=27)  # 28 giorni incluso oggi

    data = _load_log(config)
    sessioni = data.get("sessioni", [])

    # Anche archivio se presente
    try:
        arch_path = Path(config.get("log_path", "")).parent / "lavorazioni_log_archivio.json"
        if arch_path.exists():
            arch = json.loads(arch_path.read_text(encoding="utf-8"))
            sessioni = sessioni + arch.get("sessioni", [])
    except Exception:
        pass

    # Griglia: {data_iso: {fascia_key: sec}}
    griglia: dict[str, dict[str, int]] = {}
    for i in range(28):
        d = (da_data + timedelta(days=i)).isoformat()
        griglia[d] = {"mattina": 0, "sera": 0, "notte": 0}

    def minuti_del_giorno(ts_iso: str) -> int:
        """Converte timestamp ISO in minuti dall'inizio del giorno."""
        try:
            t = dt.fromisoformat(ts_iso)
            return t.hour * 60 + t.minute + t.second // 60
        except Exception:
            return 0

    def secondi_in_fascia(inizio_iso: str, fine_iso: str, fascia: dict) -> int:
        """
        Calcola quanti secondi di una sessione cadono in una fascia oraria.
        Gestisce la fascia notte che attraversa la mezzanotte.
        """
        try:
            t_in = dt.fromisoformat(inizio_iso)
            t_fin = dt.fromisoformat(fine_iso)
        except Exception:
            return 0

        tot = 0
        # Spezza la sessione in slot da 1 minuto e conta quanti cadono in fascia
        # Approssimazione efficiente: usa i bordi della fascia
        fa, fb = fascia["da"], fascia["a"]  # minuti

        # Itera su ogni minuto della sessione (cap a 16h per sicurezza)
        durata_min = int((t_fin - t_in).total_seconds() / 60)
        durata_min = min(durata_min, 16 * 60)

        for m in range(durata_min):
            ts = t_in + timedelta(minutes=m)
            mdg = ts.hour * 60 + ts.minute

            if fascia["key"] == "notte":
                # Notte: 22:00-24:00 oppure 00:00-07:30
                in_fascia = mdg >= 22*60 or mdg < 7*60+30
            else:
                in_fascia = fa <= mdg < fb

            if in_fascia:
                tot += 60  # 1 minuto = 60 secondi

        return tot

    for sess in sessioni:
        inizio = sess.get("inizio")
        fine   = sess.get("fine")
        if not inizio or not fine:
            continue
        try:
            data_sess = dt.fromisoformat(inizio).date()
        except Exception:
            continue
        if data_sess < da_data or data_sess > oggi:
            continue

        data_iso = data_sess.isoformat()
        if data_iso not in griglia:
            continue

        for fascia in FASCE:
            sec = secondi_in_fascia(inizio, fine, fascia)
            griglia[data_iso][fascia["key"]] += sec

    # Formatta output
    giorni = []
    for i in range(28):
        d = da_data + timedelta(days=i)
        d_iso = d.isoformat()
        celle = griglia[d_iso]
        giorni.append({
            "data":    d_iso,
            "giorno":  d.strftime("%a"),   # Lun, Mar, ...
            "dom":     d.day,
            "mese":    d.month,
            "celle":   {
                k: {
                    "sec":  v,
                    "ore":  round(v / 3600, 1),
                    "pct":  min(100, round(v / MAX_SEC_CELLA * 100)),
                }
                for k, v in celle.items()
            }
        })

    return {
        "giorni":      giorni,
        "fasce":       [{"key": f["key"], "label": f["label"]} for f in FASCE],
        "max_sec":     MAX_SEC_CELLA,
        "da_data":     da_data.isoformat(),
        "a_data":      oggi.isoformat(),
    }
